"""routes/task_routes.py — 任务路由

包含：
- GET  /tasks              任务列表（筛选 + 搜索 + 排序 + 分页）
- GET  /tasks/new          显示新建任务表单
- POST /tasks/new          创建任务
- GET  /tasks/<id>         任务详情独立页（V2 保留：消息跳转 / JS 失效降级 / 浏览器直链）
- GET  /tasks/<id>/drawer  任务详情抽屉 HTML 片段（V2 新增，Q3）
- POST /tasks/<id>/field   行内编辑单字段保存（V2 新增，JSON）
- POST /tasks/<id>/evidence    添加过程证据（V2 新增，JSON）
- POST /tasks/<id>/blockers    添加阻塞记录（V2 新增，JSON）
- POST /tasks/<id>/blockers/<bid>/resolve  标记阻塞解决（V2 新增，JSON）
- POST /tasks/<id>/evidence/<eid>/delete   删除过程证据（V2 批次 4 新增，仅 admin，AJAX JSON / 普通 POST 重定向）
- POST /tasks/<id>/blockers/<bid>/delete   删除阻塞记录（V2 批次 4 新增，仅 admin，AJAX JSON / 普通 POST 重定向）
- GET  /tasks/owner/<uid>/drawer  Owner 抽屉 HTML 片段（V2 新增，Q8）
- POST /tasks/remind       推送提醒（V2 新增，admin 发站内信，Q8）
- POST /tasks/<id>/edit    编辑任务字段
- POST /tasks/<id>/status  变更任务状态（V2：AJAX 请求返回 JSON）
- POST /tasks/<id>/delete  物理删除任务（仅管理员，仅已撤销任务）
- GET  /tasks/export       导出当前筛选条件下的任务为 Excel(.xlsx)（P1-004，每列自适应列宽）
- POST /tasks/batch        批量操作任务（批量改状态 / 批量指派，P1-003）
"""

import math
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, abort, Response

import config
import mail_dispatcher
import models
import auth
from auth import login_required, admin_required, get_current_user, can_edit_task, can_delete_task
import state_machine
from state_machine import (
    TaskStatus, STATUS_LABELS, STATUS_COLORS,
    TaskPriority, PRIORITY_LABELS, ALL_STATUSES, ALL_PRIORITIES,
    change_task_status, get_allowed_transitions,
)

task_bp = Blueprint('task', __name__)


@task_bp.route('/tasks')
@login_required
def task_list():
    """任务列表页：支持筛选、搜索、排序、分页（C1 确认）。"""
    user = get_current_user()

    # 解析查询参数
    filters = {
        'status':   request.args.get('status', ''),
        'priority': request.args.get('priority', ''),
        'assignee': request.args.get('assignee', ''),
        'keyword':  request.args.get('keyword', ''),
        'sort':     request.args.get('sort', 'due_date_asc'),
    }
    # V2：兼容 owner= 参数（闭环矩阵负责人名字跳转，映射到 assignee 筛选）
    owner_param = request.args.get('owner', '')
    if owner_param and not filters['assignee']:
        filters['assignee'] = owner_param
    # 清理空值
    filters = {k: v for k, v in filters.items() if v}

    # 分页参数
    try:
        page = int(request.args.get('page', 1))
    except (ValueError, TypeError):
        page = 1
    if page < 1:
        page = 1

    # 查询任务
    tasks, total = models.get_tasks(filters, page=page, per_page=config.TASKS_PER_PAGE)
    total_pages = max(1, math.ceil(total / config.TASKS_PER_PAGE))

    # 获取启用的用户列表（负责人下拉选项）
    active_users = models.get_all_active_users()

    return render_template(
        'tasks/list.html',
        tasks=tasks,
        filters=filters,
        page=page,
        total_pages=total_pages,
        total=total,
        per_page=config.TASKS_PER_PAGE,
        active_users=active_users,
        statuses=ALL_STATUSES,
        status_labels=STATUS_LABELS,
        status_colors=STATUS_COLORS,
        priorities=ALL_PRIORITIES,
        priority_labels=PRIORITY_LABELS,
        current_user=user,
        can_edit_task=can_edit_task,
    )


@task_bp.route('/tasks/new', methods=['GET', 'POST'])
@login_required
def task_new():
    """新建任务：GET 显示表单，POST 创建。"""
    user = get_current_user()

    if request.method == 'POST':
        title       = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        assignee    = request.form.get('assignee', '').strip()
        priority    = request.form.get('priority', 'medium')
        due_date    = request.form.get('due_date', '')

        # V2 批次 4：新增 3 字段（进度 / 风险点 / 协同方）
        progress_percent, risk_note, collaborators, extra_errors = _parse_form_extra_fields()

        # --- 表单校验 ---
        errors = list(extra_errors)
        if not title:
            errors.append('请输入任务标题')
        elif len(title) > 100:
            errors.append('任务标题不能超过 100 字')

        if not due_date:
            errors.append('请选择截止日期')

        # owner 只能指派给自己
        if user['role'] == 'owner':
            assignee = str(user['user_id'])
        elif not assignee:
            errors.append('请选择负责人')
        else:
            # 校验负责人是否有效
            assignee_user = models.get_user(int(assignee))
            if not assignee_user:
                errors.append('所选负责人不存在')

        if priority not in ALL_PRIORITIES:
            priority = 'medium'

        if errors:
            for err in errors:
                flash(err, 'error')
            active_users = models.get_all_active_users()
            return render_template(
                'tasks/form.html',
                active_users=active_users,
                priorities=ALL_PRIORITIES,
                priority_labels=PRIORITY_LABELS,
                form_data=request.form,
                is_edit=False,
            )

        # --- 创建任务 ---
        task_id = models.create_task(
            title=title,
            description=description,
            created_by=user['user_id'],
            assignee=int(assignee),
            priority=priority,
            due_date=due_date,
        )

        # V2 批次 4：保存新增 3 字段（不改 create_task 旧签名，创建后补写）
        if progress_percent or risk_note or collaborators:
            models.update_task(
                task_id,
                progress_percent=progress_percent,
                risk_note=risk_note,
                collaborators=collaborators,
            )

        # 生成任务指派消息通知负责人（如果负责人不是创建人自己）
        if int(assignee) != user['user_id']:
            _notify_assignment(
                task_id, user, int(assignee),
                f'{user["display_name"]} 给你指派了新任务「{title}」',
            )

        flash(f'任务「{title}」创建成功', 'success')
        # 跳转到任务列表（而非任务详情页），避免直接访问 detail 模板（侧滑片段）出现的丑陋界面
        return redirect(url_for('task.task_list'))

    # GET：显示新建表单
    active_users = models.get_all_active_users()
    return render_template(
        'tasks/form.html',
        active_users=active_users,
        priorities=ALL_PRIORITIES,
        priority_labels=PRIORITY_LABELS,
        form_data={},
        is_edit=False,
        current_user=user,
    )


@task_bp.route('/tasks/<int:task_id>')
@login_required
def task_detail(task_id):
    """任务详情（侧滑面板内容）。"""
    user = get_current_user()
    task = models.get_task(task_id)
    if not task:
        abort(404)

    # 获取进度记录时间线
    logs = models.get_progress_logs(task_id)

    # V2：过程证据 + 阻塞记录（独立详情页只读展示）
    evidence = models.get_evidence_list(task_id)
    blockers = models.get_blockers(task_id)

    # 获取可转换的状态列表
    allowed_transitions = []
    if can_edit_task(task, user):
        allowed_transitions = get_allowed_transitions(task['status'], user['role'])

    # 获取启用的用户列表（负责人下拉，仅管理员可改负责人）
    active_users = models.get_all_active_users()

    # V5 Phase 1 ②：详情页 AI 催办话术入口上下文
    ai_enabled = bool(config.AI_ENABLED)
    can_use_ai = user['role'] == 'admin' and ai_enabled
    ai_log_id = request.args.get('ai_log_id', type=int)
    ai_log = models.get_ai_log(ai_log_id) if ai_log_id else None

    return render_template(
        'tasks/detail.html',
        task=task,
        logs=logs,
        evidence=evidence,
        blockers=blockers,
        allowed_transitions=allowed_transitions,
        active_users=active_users,
        statuses=ALL_STATUSES,
        status_labels=STATUS_LABELS,
        status_colors=STATUS_COLORS,
        priorities=ALL_PRIORITIES,
        priority_labels=PRIORITY_LABELS,
        evidence_type_labels=models.EVIDENCE_TYPE_LABELS,
        blocker_status_labels=models.BLOCKER_STATUS_LABELS,
        current_user=user,
        can_edit=can_edit_task(task, user),
        can_delete=can_delete_task(task, user),
        ai_enabled=ai_enabled,
        can_use_ai=can_use_ai,
        ai_log=ai_log,
        **_mail_button_context(task, user),
    )


@task_bp.route('/tasks/<int:task_id>/edit', methods=['POST'])
@login_required
def task_edit(task_id):
    """编辑任务字段（owner 仅自己的任务）。"""
    user = get_current_user()
    task = models.get_task(task_id)
    if not task:
        abort(404)

    # 权限校验：owner 只能编辑自己负责的任务
    if not can_edit_task(task, user):
        flash('您无权编辑此任务', 'error')
        return redirect(url_for('task.task_list'))

    title       = request.form.get('title', '').strip()
    description = request.form.get('description', '').strip()
    priority    = request.form.get('priority', task['priority'])
    due_date    = request.form.get('due_date', task['due_date'])

    # owner 不能修改负责人
    if user['role'] == 'admin':
        assignee = request.form.get('assignee', str(task['assignee']))
    else:
        assignee = str(task['assignee'])

    # V2 批次 4：新增 3 字段（进度 / 风险点 / 协同方），校验规则与 task_field 一致。
    # 兼容旧客户端：表单未携带这 3 个字段时保留原值（避免老表单提交把字段清零）
    if ('progress_percent' in request.form or 'risk_note' in request.form
            or 'collaborators' in request.form):
        progress_percent, risk_note, collaborators, extra_errors = _parse_form_extra_fields()
    else:
        progress_percent = task['progress_percent'] or 0
        risk_note = task['risk_note'] or ''
        collaborators = task['collaborators'] or ''
        extra_errors = []

    # 校验
    if not title:
        flash('请输入任务标题', 'error')
        return redirect(url_for('task.task_detail', task_id=task_id))
    if extra_errors:
        for err in extra_errors:
            flash(err, 'error')
        return redirect(url_for('task.task_detail', task_id=task_id))

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    models.update_task(
        task_id,
        title=title,
        description=description,
        assignee=int(assignee),
        priority=priority,
        due_date=due_date,
        progress_percent=progress_percent,
        risk_note=risk_note,
        collaborators=collaborators,
        updated_at=now,
    )

    flash('任务已更新', 'success')
    # 跳转到任务列表（而非任务详情页）
    return redirect(url_for('task.task_list'))


@task_bp.route('/tasks/<int:task_id>/status', methods=['POST'])
@login_required
def task_change_status(task_id):
    """变更任务状态（含进度备注 + 副作用）。"""
    user = get_current_user()
    task = models.get_task(task_id)
    if not task:
        abort(404)

    # 权限校验
    if not can_edit_task(task, user):
        flash('您无权变更此任务状态', 'error')
        return redirect(url_for('task.task_list'))

    target_status = request.form.get('status', '')
    progress_note = request.form.get('progress_note', '').strip()

    # 闭环时需要完成说明（可选，但鼓励填写）
    if target_status == TaskStatus.CLOSED and not progress_note:
        # 不强制，但提示
        pass

    ok, reason = change_task_status(
        task_id=task_id,
        target_status=target_status,
        operator_id=user['user_id'],
        operator_role=user['role'],
        progress_note=progress_note,
    )

    # V2：抽屉内 AJAX 提交（drawer-form）时返回 JSON，前端 reload 抽屉
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        if ok:
            return jsonify({
                'success': True,
                'message': '任务状态已更新为「{}」'.format(STATUS_LABELS.get(target_status, target_status)),
            })
        return jsonify({'success': False, 'message': reason})

    if ok:
        flash(f'任务状态已更新为「{STATUS_LABELS.get(target_status, target_status)}」', 'success')
    else:
        flash(reason, 'error')

    return redirect(url_for('task.task_detail', task_id=task_id))


@task_bp.route('/tasks/<int:task_id>/delete', methods=['POST'])
@admin_required
def task_delete(task_id):
    """物理删除任务（仅管理员，仅允许删除已撤销任务）。"""
    user = get_current_user()
    task = models.get_task(task_id)
    if not task:
        abort(404)

    # 仅允许删除已撤销的任务（PRD Q6-B）
    if task['status'] != 'cancelled':
        flash('仅允许删除已撤销的任务', 'error')
        return redirect(url_for('task.task_list'))

    models.delete_task(task_id)
    flash('任务已永久删除', 'success')
    return redirect(url_for('task.task_list'))


# ============================================================
# P1-004：任务导出 Excel(.xlsx)
# ============================================================

@task_bp.route('/tasks/export')
@login_required
def task_export():
    """导出当前筛选条件下的任务列表为 Excel(.xlsx) 文件（P1-004）。

    使用 openpyxl 生成 .xlsx，支持「每列自适应列宽」（按本列最长内容自动撑开列宽），
    表头加粗 + 浅蓝底、首行冻结、启用自动筛选，Excel / WPS 打开即整齐可读。
    沿用 /tasks 的筛选参数（status/priority/assignee/keyword/sort），导出当前筛选条件下的全部任务（不分页）。

    Returns:
        Response: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet 附件，
                  文件名 督办任务列表_YYYY-MM-DD.xlsx
    """
    # 解析筛选参数（与 task_list 完全一致）
    filters = {
        'status':   request.args.get('status', ''),
        'priority': request.args.get('priority', ''),
        'assignee': request.args.get('assignee', ''),
        'keyword':  request.args.get('keyword', ''),
        'sort':     request.args.get('sort', 'due_date_asc'),
    }
    filters = {k: v for k, v in filters.items() if v}

    # 取全部任务（不分页，per_page 取一个大数）
    tasks, _ = models.get_tasks(filters, page=1, per_page=100000)

    # 表头（中文）与数据行
    headers = [
        '任务ID', '标题', '负责人', '创建人', '优先级', '状态',
        '截止日期', '创建时间', '更新时间', '闭环时间', '是否逾期'
    ]
    rows = []
    for task in tasks:
        rows.append([
            task['task_id'],
            task['title'],
            task['assignee_name'] or '',
            task['creator_name'] or '',
            PRIORITY_LABELS.get(task['priority'], task['priority']),
            STATUS_LABELS.get(task['status'], task['status']),
            task['due_date'] or '',
            task['created_at'] or '',
            task['updated_at'] or '',
            task['closed_at'] or '',
            '是' if task['is_overdue'] else '否',
        ])

    # 生成工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = '督办任务列表'
    ws.append(headers)
    for r in rows:
        ws.append(r)

    # 表头样式：加粗 + 浅蓝底 + 居中
    header_fill = PatternFill('solid', fgColor='DDEBF7')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 每列自适应列宽：按本列最长「显示宽度」撑开，CJK/全角字符约按 1.8 宽计，留 2 字符余量
    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        max_w = 0
        for cell in ws[letter]:
            if cell.value is None:
                continue
            txt = str(cell.value)
            # 估算显示宽度：CJK/全角字符约 1.8，其余（ASCII/数字）约 1.0
            w = sum(1.8 if ord(ch) > 0x2E80 else 1.0 for ch in txt)
            if w > max_w:
                max_w = w
        ws.column_dimensions[letter].width = max(8.0, min(60.0, max_w + 2.0))

    # 首行冻结 + 自动筛选
    ws.freeze_panes = 'A2'
    if ws.max_row >= 1:
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f'A1:{last_col}{ws.max_row}'

    # 输出为二进制（xlsx 是 zip 包，内部已用 UTF-8，无需 BOM）
    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    date_str = datetime.now().strftime('%Y-%m-%d')
    # 文件名双 fallback（RFC 5987）：
    # - 旧浏览器/纯 ASCII 客户端用 filename="tasks_YYYY-MM-DD.xlsx"
    # - 现代浏览器用 filename*=UTF-8''... 看中文名"督办任务列表_YYYY-MM-DD.xlsx"
    # 注意：HTTP header 不支持中文裸传（Werkzeug dev server 用 latin-1 strict 编码，会抛 ValueError）
    ascii_filename = f'tasks_{date_str}.xlsx'
    utf8_filename = f'督办任务列表_{date_str}.xlsx'
    from urllib.parse import quote
    filename_star = f"UTF-8''{quote(utf8_filename)}"
    response = Response(
        xlsx_bytes,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename="{ascii_filename}"; filename*={filename_star}'
        }
    )
    return response


# ============================================================
# P1-003：任务批量操作（批量改状态 + 批量指派）
# ============================================================

@task_bp.route('/tasks/batch', methods=['POST'])
@login_required
def task_batch():
    """任务批量操作（P1-003）。

    支持两种动作：
    - action=change_status：批量变更任务状态（逐个调用 state_machine.change_task_status，复用副作用：写进度日志 + 发消息）
    - action=reassign：批量指派负责人（仅管理员）

    权限规则：
    - admin：可批量操作所有任务
    - owner：只能批量操作自己负责的任务（不属于自己的跳过并计入失败数）

    成功后 flash 提示「已批量更新 N 条任务」，失败数也在 flash 中说明。

    Args (form):
        task_ids: 多个任务ID（form list，name=task_ids）
        action: change_status / reassign
        target_status: 当 action=change_status 时为新状态
        new_assignee: 当 action=reassign 时为新负责人 user_id

    Returns:
        redirect: 回任务列表页（带原筛选参数）
    """
    user = get_current_user()

    # 取选中的任务 ID 列表
    task_id_strs = request.form.getlist('task_ids')
    action = request.form.get('action', '').strip()

    # 携带原筛选参数用于回跳
    redirect_params = {
        'status':   request.form.get('filter_status', ''),
        'priority': request.form.get('filter_priority', ''),
        'assignee': request.form.get('filter_assignee', ''),
        'keyword':  request.form.get('filter_keyword', ''),
        'sort':     request.form.get('filter_sort', 'due_date_asc'),
    }
    redirect_params = {k: v for k, v in redirect_params.items() if v}

    # --- 参数校验 ---
    if not task_id_strs:
        flash('请至少选择一条任务', 'error')
        return redirect(url_for('task.task_list', **redirect_params))

    if action not in ('change_status', 'reassign'):
        flash('未知的批量操作类型', 'error')
        return redirect(url_for('task.task_list', **redirect_params))

    # reassign 仅管理员可用
    if action == 'reassign' and user['role'] != 'admin':
        flash('批量指派仅管理员可操作', 'error')
        return redirect(url_for('task.task_list', **redirect_params))

    # 校验目标值
    target_status = request.form.get('target_status', '').strip()
    new_assignee = request.form.get('new_assignee', '').strip()

    if action == 'change_status':
        if target_status not in ALL_STATUSES:
            flash('请选择有效的目标状态', 'error')
            return redirect(url_for('task.task_list', **redirect_params))
    else:  # reassign
        if not new_assignee:
            flash('请选择新负责人', 'error')
            return redirect(url_for('task.task_list', **redirect_params))
        new_assignee_user = models.get_user(int(new_assignee))
        if not new_assignee_user or not new_assignee_user['is_active']:
            flash('所选新负责人不存在或已停用', 'error')
            return redirect(url_for('task.task_list', **redirect_params))

    # --- 逐条执行（复用现有副作用逻辑，不直接 UPDATE）---
    success_count = 0
    fail_count = 0
    fail_reasons = []

    for tid_str in task_id_strs:
        try:
            tid = int(tid_str)
        except (ValueError, TypeError):
            fail_count += 1
            fail_reasons.append(f'无效的任务ID: {tid_str}')
            continue

        task = models.get_task(tid)
        if not task:
            fail_count += 1
            fail_reasons.append(f'任务 {tid} 不存在')
            continue

        # 权限校验：owner 只能操作自己的任务
        if not can_edit_task(task, user):
            fail_count += 1
            fail_reasons.append(f'无权操作任务「{task["title"]}」')
            continue

        if action == 'change_status':
            # 调用状态机（复用副作用：写进度日志 + 发消息）
            ok, reason = change_task_status(
                task_id=tid,
                target_status=target_status,
                operator_id=user['user_id'],
                operator_role=user['role'],
                progress_note='批量操作：变更为「{}」'.format(STATUS_LABELS.get(target_status, target_status)),
            )
            if ok:
                success_count += 1
            else:
                fail_count += 1
                fail_reasons.append(f'「{task["title"]}」: {reason}')

        else:  # reassign（仅 admin，已在上面校验）
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            models.update_task(
                tid,
                assignee=int(new_assignee),
                updated_at=now,
            )
            # 生成指派消息通知新负责人（如果新负责人不是操作人自己）
            if int(new_assignee) != user['user_id']:
                _notify_assignment(
                    tid, user, int(new_assignee),
                    '{} 通过批量指派把任务「{}」转交给你'.format(user['display_name'], task['title']),
                )
            success_count += 1

    # --- 结果反馈 ---
    if success_count and not fail_count:
        flash('已批量更新 {} 条任务'.format(success_count), 'success')
    elif success_count and fail_count:
        flash('已批量更新 {} 条，{} 条失败：{}'.format(
            success_count, fail_count, '；'.join(fail_reasons[:3])
        ), 'warning')
    else:
        flash('批量操作失败：{}'.format('；'.join(fail_reasons[:3])), 'error')

    return redirect(url_for('task.task_list', **redirect_params))


# ============================================================
# V2 批次 3：任务详情抽屉 + 行内编辑 + 证据/阻塞 + Owner 抽屉 + 推送提醒
# ============================================================

# 行内编辑字段白名单（§5.1）
FIELD_WHITELIST = (
    'title', 'description', 'priority', 'due_date',
    'progress_percent', 'risk_note', 'collaborators', 'assignee',
)

_DATE_RE = re.compile(r'^\d{4}-\d{2}-\d{2}$')

# V2 批次 4：risk_note / collaborators 文本长度上限（对齐证据/阻塞内容的 500 字上限）
TEXT_FIELD_MAX_LEN = 500


def _parse_form_extra_fields():
    """解析并校验表单中的 V2 新增 3 字段（progress_percent / risk_note / collaborators）。

    校验规则与 task_field 路由的 FIELD_WHITELIST 逻辑保持一致：
    - progress_percent：整数 0-100，默认 0
    - risk_note / collaborators：自由文本，去首尾空白，上限 500 字

    Returns:
        tuple: (progress_percent: int, risk_note: str, collaborators: str, errors: list[str])
    """
    errors = []

    # --- 进度（%） ---
    raw_progress = (request.form.get('progress_percent', '') or '').strip()
    if raw_progress == '':
        raw_progress = '0'
    progress_percent = 0
    try:
        progress_percent = int(raw_progress)
    except (ValueError, TypeError):
        errors.append('进度必须是 0-100 的整数')
    else:
        if progress_percent < 0 or progress_percent > 100:
            errors.append('进度必须在 0-100 之间')

    # --- 风险点说明 / 协同方 ---
    risk_note = (request.form.get('risk_note', '') or '').strip()
    collaborators = (request.form.get('collaborators', '') or '').strip()
    if len(risk_note) > TEXT_FIELD_MAX_LEN:
        errors.append(f'风险点说明不能超过 {TEXT_FIELD_MAX_LEN} 字')
    if len(collaborators) > TEXT_FIELD_MAX_LEN:
        errors.append(f'协同方不能超过 {TEXT_FIELD_MAX_LEN} 字')

    return progress_percent, risk_note, collaborators, errors


@task_bp.route('/tasks/<int:task_id>/drawer')
@login_required
def task_drawer(task_id):
    """任务详情抽屉 HTML 片段（Q3 核心，3 页签：详情/过程证据/阻塞记录）。

    返回纯片段（无 base.html 包裹），由 main.js 的 drawer 管理器注入抽屉容器。
    行内编辑权限：admin 全部可改；owner 只能改自己的任务（他人任务字段只读展示）。
    """
    user = get_current_user()
    task = models.get_task(task_id)
    if not task:
        abort(404)

    logs = models.get_progress_logs(task_id)
    evidence = models.get_evidence_list(task_id)
    blockers = models.get_blockers(task_id)

    allowed_transitions = []
    if can_edit_task(task, user):
        allowed_transitions = get_allowed_transitions(task['status'], user['role'])

    active_users = models.get_all_active_users()

    return render_template(
        'tasks/_drawer.html',
        task=task,
        logs=logs,
        evidence=evidence,
        blockers=blockers,
        allowed_transitions=allowed_transitions,
        active_users=active_users,
        status_labels=STATUS_LABELS,
        priorities=ALL_PRIORITIES,
        priority_labels=PRIORITY_LABELS,
        evidence_type_labels=models.EVIDENCE_TYPE_LABELS,
        blocker_status_labels=models.BLOCKER_STATUS_LABELS,
        current_user=user,
        can_edit=can_edit_task(task, user),
        can_delete=can_delete_task(task, user),
        **_mail_button_context(task, user),
    )


@task_bp.route('/tasks/<int:task_id>/field', methods=['POST'])
@login_required
def task_field(task_id):
    """行内编辑单字段保存（JSON 请求/响应）。

    请求体：{"field": "title", "value": "新标题"}
    权限：can_edit_task（admin 全部；owner 仅自己的任务；assignee 仅 admin 可改）。
    """
    user = get_current_user()
    task = models.get_task(task_id)
    if not task:
        return jsonify({'success': False, 'message': '任务不存在'}), 404

    if not can_edit_task(task, user):
        return jsonify({'success': False, 'message': '您无权编辑此任务'}), 403

    data = request.get_json(silent=True) or {}
    field = data.get('field', '')
    value = data.get('value')

    # --- 白名单校验 ---
    if field not in FIELD_WHITELIST:
        return jsonify({'success': False, 'message': '不支持的编辑字段'}), 400

    # --- 逐字段校验与规范化 ---
    if field == 'title':
        value = (value or '').strip()
        if not value:
            return jsonify({'success': False, 'message': '任务标题不能为空'}), 400
        if len(value) > 100:
            return jsonify({'success': False, 'message': '任务标题不能超过 100 字'}), 400

    elif field == 'priority':
        if value not in ALL_PRIORITIES:
            return jsonify({'success': False, 'message': '无效的优先级'}), 400

    elif field == 'due_date':
        value = (value or '').strip()
        if not _DATE_RE.match(value):
            return jsonify({'success': False, 'message': '请输入有效的截止日期（YYYY-MM-DD）'}), 400

    elif field == 'progress_percent':
        try:
            value = int(value)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': '进度必须是 0-100 的整数'}), 400
        if value < 0 or value > 100:
            return jsonify({'success': False, 'message': '进度必须在 0-100 之间'}), 400

    elif field == 'assignee':
        # 负责人变更仅管理员可操作
        if user['role'] != 'admin':
            return jsonify({'success': False, 'message': '仅管理员可修改负责人'}), 403
        try:
            value = int(value)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': '无效的负责人'}), 400
        assignee_user = models.get_user(value)
        if not assignee_user or not assignee_user['is_active']:
            return jsonify({'success': False, 'message': '所选负责人不存在或已停用'}), 400

    else:
        # description / risk_note / collaborators：自由文本
        value = (value or '').strip()

    # --- 保存 ---
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    models.update_task(task_id, updated_at=now, **{field: value})

    # 负责人变更时给新负责人发通知
    if field == 'assignee' and value != task['assignee'] and value != user['user_id']:
        _notify_assignment(
            task_id, user, value,
            '{} 把任务「{}」转交给你'.format(user['display_name'], task['title']),
        )

    result = {'success': True, 'message': '已保存'}
    if field == 'progress_percent':
        result['progress'] = value
    return jsonify(result)


@task_bp.route('/tasks/<int:task_id>/evidence', methods=['POST'])
@login_required
def task_evidence_add(task_id):
    """添加过程证据（抽屉内表单 AJAX 提交，JSON 响应）。"""
    user = get_current_user()
    task = models.get_task(task_id)
    if not task:
        return jsonify({'success': False, 'message': '任务不存在'}), 404

    if not can_edit_task(task, user):
        return jsonify({'success': False, 'message': '您无权添加证据'}), 403

    etype = request.form.get('etype', '').strip()
    content = request.form.get('content', '').strip()

    if etype not in models.EVIDENCE_TYPES:
        return jsonify({'success': False, 'message': '请选择有效的证据类型'}), 400
    if not content:
        return jsonify({'success': False, 'message': '请输入证据内容'}), 400
    if len(content) > 500:
        return jsonify({'success': False, 'message': '证据内容不能超过 500 字'}), 400

    models.add_evidence(task_id, etype, content, user['user_id'])

    # 时间线留痕（设计文档 §4.4：所有证据/阻塞操作写进 progress_logs）
    _log_timeline(task_id, user['user_id'],
                  '添加证据（{}）：{}'.format(models.EVIDENCE_TYPE_LABELS.get(etype, etype),
                                              _brief(content)))
    return jsonify({'success': True, 'message': '证据已添加'})


@task_bp.route('/tasks/<int:task_id>/blockers', methods=['POST'])
@login_required
def task_blocker_add(task_id):
    """添加阻塞记录（抽屉内表单 AJAX 提交，JSON 响应）。"""
    user = get_current_user()
    task = models.get_task(task_id)
    if not task:
        return jsonify({'success': False, 'message': '任务不存在'}), 404

    if not can_edit_task(task, user):
        return jsonify({'success': False, 'message': '您无权添加阻塞记录'}), 403

    content = request.form.get('content', '').strip()
    if not content:
        return jsonify({'success': False, 'message': '请输入阻塞描述'}), 400
    if len(content) > 500:
        return jsonify({'success': False, 'message': '阻塞描述不能超过 500 字'}), 400

    models.add_blocker(task_id, content, user['user_id'])

    # 时间线留痕
    _log_timeline(task_id, user['user_id'], '添加阻塞：{}'.format(_brief(content)))
    return jsonify({'success': True, 'message': '阻塞记录已添加'})


@task_bp.route('/tasks/<int:task_id>/blockers/<int:blocker_id>/resolve', methods=['POST'])
@login_required
def task_blocker_resolve(task_id, blocker_id):
    """标记阻塞记录为已解决（JSON 响应）。

    权限：admin 可解决任何阻塞；owner 仅能解决自己创建的阻塞记录。
    """
    user = get_current_user()
    task = models.get_task(task_id)
    if not task:
        return jsonify({'success': False, 'message': '任务不存在'}), 404

    blocker = db_blocker_or_none(blocker_id, task_id)
    if blocker is None:
        return jsonify({'success': False, 'message': '阻塞记录不存在'}), 404

    is_admin = (user['role'] == 'admin')
    is_creator = (blocker['created_by'] == user['user_id'])
    if not (is_admin or is_creator):
        return jsonify({'success': False, 'message': '仅管理员或记录创建者可标记解决'}), 403

    ok = models.resolve_blocker(blocker_id, user['user_id'])
    if not ok:
        return jsonify({'success': False, 'message': '该阻塞记录已是已解决状态'}), 400

    # 时间线留痕
    _log_timeline(task_id, user['user_id'],
                  '标记阻塞解决：{}'.format(_brief(blocker['content'])))
    return jsonify({'success': True, 'message': '已标记为解决'})


def db_blocker_or_none(blocker_id, task_id):
    """查询属于指定任务的阻塞记录（不存在或不属于该任务返回 None）。"""
    from models import db
    return db.query_one(
        "SELECT * FROM blockers WHERE blocker_id = ? AND task_id = ?",
        (blocker_id, task_id)
    )


def db_evidence_or_none(evidence_id, task_id):
    """查询属于指定任务的过程证据（不存在或不属于该任务返回 None）。"""
    from models import db
    return db.query_one(
        "SELECT * FROM evidence WHERE evidence_id = ? AND task_id = ?",
        (evidence_id, task_id)
    )


def _brief(text, limit=50):
    """截断文本用于时间线备注（超长加省略号，保持时间线整洁）。"""
    text = (text or '').strip()
    if len(text) <= limit:
        return text
    return text[:limit] + '…'


def _mail_button_context(task, user):
    """算出「手动发送提醒邮件」按钮的渲染参数（抽屉与整页详情共用）。

    未配置邮件时按钮仍然渲染，只是换成说明文案——对操作人来说，
    看见「有这个功能、但要管理员先配置」比压根没有这个入口更有用。

    Returns:
        dict: can_send_mail / mail_ready / mail_cooling / mail_cooldown_min
    """
    cfg = models.get_mail_config()
    mail_ready = models.is_mail_configured(cfg)

    # H1-②：管理员 + 任务创建人。负责人给自己发提醒没有意义，
    # 但「我布置的任务我来催」是最常见场景，只给管理员会绕远路。
    can_send_mail = user['role'] == 'admin' or task['created_by'] == user['user_id']

    cooldown = int(cfg.get('manual_cooldown') or 0)
    mail_cooling = bool(
        can_send_mail and mail_ready and cooldown > 0
        and models.has_recent_manual_mail(task['task_id'], user['user_id'], cooldown)
    )

    return {
        'can_send_mail': can_send_mail,
        'mail_ready': mail_ready,
        'mail_cooling': mail_cooling,
        'mail_cooldown_min': max(1, round(cooldown / 60)) if cooldown else 0,
    }


def _notify_assignment(task_id, operator, assignee_id, message):
    """任务分配 / 改派后的通知：站内信 + 邮件双通道。

    站内信照旧由调用方写好文案经本函数发出；邮件则在函数内部
    委托 mail_dispatcher 入队（C5-①：分配是任务生命周期的起点，
    负责人不上系统看就永远不知道自己多了个任务）。

    邮件是**附加通道**，任何异常都不允许影响站内信与业务主流程——
    所以这里单独 try/except 吞掉并打印，与 task_routes 既有的
    「留痕失败不阻断主操作」风格保持一致。

    Args:
        task_id: 任务 ID
        operator: 操作人用户行（需含 user_id / display_name / email）
        assignee_id: 新负责人 user_id
        message: 站内信正文
    """
    models.create_message(
        recipient=assignee_id,
        sender=operator['user_id'],
        msg_type='assignment',
        content=message,
        task_id=task_id,
    )

    try:
        mail_dispatcher.enqueue_assignment(
            task_id,
            operator_id=operator['user_id'],
            is_transfer=True,
            operator_email=(operator['email'] or None) if 'email' in operator.keys() else None,
        )
    except Exception as e:  # pragma: no cover - 邮件失败不影响业务
        print(f'[mail] 分配通知邮件入队失败: {e}')


def _log_timeline(task_id, operator, note):
    """写一条纯备注型进度日志（status_from/status_to 均为 None，时间线只显示操作人与备注）。

    用于证据/阻塞操作留痕（设计文档 §4.4）。异常不阻断主流程。
    """
    try:
        models.create_progress_log(
            task_id,
            operator,
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            None,   # status_from：非状态变更
            None,   # status_to：非状态变更
            note,
        )
    except Exception as e:  # pragma: no cover - 留痕失败不影响主操作
        print(f'[timeline] 留痕写入失败: {e}')


# ============================================================
# V2 批次 4：删除证据 / 删除阻塞（仅管理员）
# ============================================================

@task_bp.route('/tasks/<int:task_id>/evidence/<int:evidence_id>/delete', methods=['POST'])
@admin_required
def task_evidence_delete(task_id, evidence_id):
    """删除过程证据（V2 批次 4，仅管理员）。

    权限规则（设计文档 §4.4）：删除仅 admin；证据必须属于该任务。
    AJAX 调用（X-Requested-With: XMLHttpRequest）返回 JSON {success, message}；
    普通 POST（详情独立页表单）重定向回任务详情页。
    """
    user = get_current_user()

    task = models.get_task(task_id)
    if not task:
        abort(404)

    evidence = db_evidence_or_none(evidence_id, task_id)
    if evidence is None:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': '证据不存在或不属于该任务'}), 404
        flash('证据不存在或不属于该任务', 'error')
        return redirect(url_for('task.task_detail', task_id=task_id))

    # 删除 + 时间线留痕
    ok = models.delete_evidence(evidence_id)
    if not ok:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': '删除失败，证据不存在'}), 404
        flash('删除失败，证据不存在', 'error')
        return redirect(url_for('task.task_detail', task_id=task_id))

    _log_timeline(task_id, user['user_id'],
                  '删除证据（{}）：{}'.format(models.EVIDENCE_TYPE_LABELS.get(evidence['etype'], evidence['etype']),
                                               _brief(evidence['content'])))

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': True, 'message': '证据已删除'})
    flash('证据已删除', 'success')
    return redirect(url_for('task.task_detail', task_id=task_id))


@task_bp.route('/tasks/<int:task_id>/blockers/<int:blocker_id>/delete', methods=['POST'])
@admin_required
def task_blocker_delete(task_id, blocker_id):
    """删除阻塞记录（V2 批次 4，仅管理员）。

    权限规则（设计文档 §4.4）：删除仅 admin；阻塞必须属于该任务。
    AJAX 调用（X-Requested-With: XMLHttpRequest）返回 JSON {success, message}；
    普通 POST（详情独立页表单）重定向回任务详情页。
    """
    user = get_current_user()

    task = models.get_task(task_id)
    if not task:
        abort(404)

    blocker = db_blocker_or_none(blocker_id, task_id)
    if blocker is None:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': '阻塞记录不存在或不属于该任务'}), 404
        flash('阻塞记录不存在或不属于该任务', 'error')
        return redirect(url_for('task.task_detail', task_id=task_id))

    # 删除 + 时间线留痕
    ok = models.delete_blocker(blocker_id)
    if not ok:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': '删除失败，阻塞记录不存在'}), 404
        flash('删除失败，阻塞记录不存在', 'error')
        return redirect(url_for('task.task_detail', task_id=task_id))

    _log_timeline(task_id, user['user_id'],
                  '删除阻塞：{}'.format(_brief(blocker['content'])))

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': True, 'message': '阻塞记录已删除'})
    flash('阻塞记录已删除', 'success')
    return redirect(url_for('task.task_detail', task_id=task_id))


@task_bp.route('/tasks/owner/<int:user_id>/drawer')
@login_required
def owner_drawer(user_id):
    """Owner 抽屉 HTML 片段（Q8）：负责人的任务列表 + 推送提醒按钮（仅 admin 可见）。

    049 增强：支持可选 ?status= 参数，按状态过滤（闭环矩阵点击某状态数字时传入）。
    status 不在白名单内则忽略（回落为全部任务）。
    """
    user = get_current_user()
    owner = models.get_user(user_id)
    if not owner:
        abort(404)

    # 可选状态过滤（仅复核白名单，避免注入/非法值）
    status_filter = request.args.get('status', '')
    if status_filter and status_filter not in ALL_STATUSES:
        status_filter = ''

    filters = {'assignee': str(user_id), 'sort': 'due_date_asc'}
    if status_filter:
        filters['status'] = status_filter

    # 该负责人最近 50 条任务（按截止日期升序）
    tasks, total = models.get_tasks(filters, page=1, per_page=50)

    return render_template(
        'tasks/_owner_drawer.html',
        owner=owner,
        tasks=tasks,
        total=total,
        status_filter=status_filter,
        status_labels=STATUS_LABELS,
        priority_labels=PRIORITY_LABELS,
        current_user=user,
    )


@task_bp.route('/tasks/remind', methods=['POST'])
@login_required
def task_remind():
    """推送提醒（Q8）：admin 给负责人发一条站内提醒消息（type=admin_directive）。

    参数（form 或 JSON）：
    - task_id: 提醒某条任务的负责人（消息中带任务标题）
    - owner_id: 提醒某位负责人（泛化提醒，与 task_id 二选一）
    """
    user = get_current_user()

    # 手动校验 admin（保证 AJAX 调用拿到 JSON 而非 302 跳转）
    if user['role'] != 'admin':
        return jsonify({'success': False, 'message': '仅管理员可推送提醒'}), 403

    # 兼容 form / JSON / query 三种参数来源
    # （remind-btn 的 data-url 把参数放在查询串里，POST 无 body）
    if request.is_json:
        params = request.get_json(silent=True) or {}
    else:
        params = {}
    params = {
        'task_id':  params.get('task_id', '') or request.form.get('task_id', '') or request.args.get('task_id', ''),
        'owner_id': params.get('owner_id', '') or request.form.get('owner_id', '') or request.args.get('owner_id', ''),
    }

    task_id = params.get('task_id', '')
    owner_id = params.get('owner_id', '')

    if task_id:
        task = models.get_task(int(task_id))
        if not task:
            return jsonify({'success': False, 'message': '任务不存在'}), 404
        recipient = task['assignee']
        content = '{} 提醒你尽快推进任务「{}」（当前状态：{}，截止 {}）'.format(
            user['display_name'], task['title'],
            STATUS_LABELS.get(task['status'], task['status']),
            task['due_date'],
        )
    elif owner_id:
        owner = models.get_user(int(owner_id))
        if not owner or not owner['is_active']:
            return jsonify({'success': False, 'message': '所选负责人不存在或已停用'}), 404
        recipient = owner['user_id']
        content = '{} 向你发送了督办提醒：请及时关注你名下任务的进展'.format(user['display_name'])
    else:
        return jsonify({'success': False, 'message': '请指定任务或负责人'}), 400

    # 不给自己发提醒
    if recipient == user['user_id']:
        return jsonify({'success': False, 'message': '不能给自己推送提醒'}), 400

    models.create_message(
        recipient=recipient,
        sender=user['user_id'],
        msg_type='admin_directive',
        content=content,
        task_id=int(task_id) if task_id else None,
    )
    return jsonify({'success': True, 'message': '已发送督办提醒'})


@task_bp.route('/tasks/<int:task_id>/send-mail', methods=['POST'])
@login_required
def task_send_mail(task_id):
    """手动发送提醒邮件（V4，C3-① 任务详情页按钮）。

    权限（H1-②）：管理员 + 该任务的创建人。
        —— 负责人不需要邮件提醒自己，逻辑上说不通；
        —— 只给管理员会让「我布置的任务我来催」这种最常见场景绕远路。

    流程：入队（拿冷却与审计）→ 立刻同步发送这一条 → 返回真实结果。
        同步发送是为了让操作人当场知道有没有发出去；
        万一同步失败也没关系，记录留在队列里由后台自动重试（G1）。

    请求参数（form 或 JSON）：
        note: 可选留言，≤200 字纯文本（H4-②）

    响应形态随调用方自动切换：
        AJAX（抽屉里的 .drawer-form）→ JSON，由 main.js 弹 Toast；
        整页表单（/tasks/<id> 详情页）→ flash + 跳回原页面，
        否则用户点完按钮会看到一屏原始 JSON。
    """
    user = get_current_user()
    task = models.get_task(task_id)
    if not task:
        return _mail_result(task_id, {'success': False, 'message': '任务不存在'}, 404)

    # --- 权限校验（H1-②）---
    is_admin = user['role'] == 'admin'
    is_creator = task['created_by'] == user['user_id']
    if not (is_admin or is_creator):
        return _mail_result(
            task_id, {'success': False, 'message': '仅管理员或任务创建人可发送提醒邮件'}, 403)

    if not models.is_mail_configured():
        return _mail_result(task_id, {
            'success': False,
            'message': '邮件功能未启用或未完成配置，请联系管理员在「邮件」页配置',
        }, 400)

    params = request.get_json(silent=True) if request.is_json else {}
    note = (params or {}).get('note', '') or request.form.get('note', '') or ''

    queue_id, err = mail_dispatcher.enqueue_manual(
        task_id,
        operator_id=user['user_id'],
        note=note,
        operator_email=(user['email'] or None) if 'email' in user.keys() else None,
    )
    if err:
        return _mail_result(task_id, {'success': False, 'message': err}, 400)

    # 入队成功 → 立刻尝试同步发送这一条（不等待 5 分钟扫描周期）
    result = mail_dispatcher.send_one(queue_id)

    assignee = models.get_user(task['assignee'])
    who = assignee['display_name'] if assignee else '负责人'

    if result['success']:
        return _mail_result(task_id, {
            'success': True,
            'message': f'提醒邮件已发送给 {who}',
        }, 200)

    return _mail_result(task_id, {
        'success': False,
        'message': '邮件已加入发送队列，但本次发送未成功，系统将在稍后自动重试。'
                   f'原因：{result.get("error_message") or "未知错误"}',
    }, 202)


def _mail_result(task_id, payload, status):
    """手动发送邮件的响应适配：AJAX 回 JSON，整页提交回 flash + 重定向。

    202（已入队待重试）在整页场景下按失败提示处理——对操作人来说
    「没发出去」才是此刻需要知道的信息，重试细节属于后台行为。
    """
    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify(payload), status

    flash(payload['message'], 'success' if payload['success'] else 'error')
    return redirect(request.referrer or url_for('task.task_detail', task_id=task_id))
