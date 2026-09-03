#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""test_suite.py — 督办系统综合测试套件

测试范围：
1. 数据库初始化测试（建表、索引、默认配置）
2. 状态机转换测试（合法/非法转换、权限校验、副作用）
3. 应用启动与路由测试（蓝图注册、路由可访问性）
4. 端到端功能测试（完整业务流程）
5. 权限测试（未登录拦截、角色隔离、数据级权限）
6. 分页测试（C1：每页 20 条）
7. 预警引擎测试（三层预警、去重）
8. V2 仪表盘测试（闭环率口径、时间范围筛选、闭环矩阵、今日焦点）
9. V2 行内编辑测试（POST /tasks/<id>/field 权限矩阵与校验）
10. V2 证据/阻塞测试（CRUD、权限、时间线留痕）
11. V2 抽屉路由测试（任务抽屉 / Owner 抽屉 / 消息抽屉）
12. V2 推送提醒测试（POST /tasks/remind）
13. V2 登录页与暗色模式测试（记住我 cookie、字段级错误、dark 模式切换）
14. V2 表单 3 字段测试（progress_percent / risk_note / collaborators）

运行方式：
    python test_suite.py
"""

import os
import sys
import math
import re
import time
import secrets
import unittest
from datetime import datetime, timedelta
from flask import request

# ============================================================
# 测试环境配置：必须在导入业务模块前设置测试数据库路径
# ============================================================
# 目录布局（src 分层后）：项目根/{src, tests, scripts, docs, data}
# 本文件位于 tests/ 下：TESTS_DIR = tests/，PROJECT_DIR = 项目根，SRC_DIR = 源码目录
TESTS_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(TESTS_DIR)
SRC_DIR = os.path.join(PROJECT_DIR, 'src')
sys.path.insert(0, SRC_DIR)

import config

# 使用独立的测试数据库，不影响生产数据
TEST_DATA_DIR = os.path.join(TESTS_DIR, 'test_data')
os.makedirs(TEST_DATA_DIR, exist_ok=True)
config.DATA_DIR = TEST_DATA_DIR
config.DATABASE_PATH = os.path.join(TEST_DATA_DIR, 'test_supervision.db')
config.DEBUG = False

# 导入业务模块
import db
import models
import auth
import csrf
import crypto_util
import state_machine
import mail_constants
import ai_service
import ai_dispatcher
from state_machine import (
    TaskStatus, STATUS_LABELS, TRANSITIONS, ADMIN_ONLY_TRANSITIONS,
    validate_transition, get_allowed_transitions, change_task_status,
)
from app import create_app


# ============================================================
# CSRF 感知的测试客户端（DEF-002）
# ============================================================

class CsrfClient:
    """自动为写请求补上 CSRF 令牌的测试客户端包装。

    为什么需要它：CSRF 防护上线后，测试里那 100 多处 .post() 全都会因为
    缺令牌返回 400。逐个加参数既费时又容易漏，而绝大多数用例关心的是
    「业务逻辑对不对」，不是「令牌有没有带」。

    做法是包一层，在真正发出请求前按**当前会话的密钥现算**一个有效令牌。
    现算这一步很关键：登录成功时后端会轮换会话密钥（rotate_csrf_token），
    如果这里用固定令牌，登录后所有写操作都会失效。

    需要**故意**测试「没带令牌会被拒」的场景时，用 client.raw 取原始客户端。
    """

    def __init__(self, client):
        self._client = client

    @property
    def raw(self):
        """未被包装的原始客户端（用于测试 CSRF 拦截本身）。"""
        return self._client

    def _token(self):
        """读出当前会话的密钥，现算一个有效令牌。"""
        with self._client.session_transaction() as sess:
            if not sess.get(csrf.SESSION_KEY):
                sess[csrf.SESSION_KEY] = secrets.token_urlsafe(32)
            secret = sess[csrf.SESSION_KEY]
        expires_at = int(time.time()) + 3600
        return f'{expires_at}.{csrf._sign(f"{secret}.{expires_at}")}'

    def _prepare(self, kwargs):
        """把令牌塞进请求。

        三种形态都覆盖到，这样套件本身也顺带验证了服务端三条取令牌的通道：
          1. 表单字段（真实浏览器走的路，也是最常见的形态）
          2. JSON body 的 csrf_token 键（行内编辑这类 JSON 接口）
          3. X-CSRF-Token 请求头（既无表单又无 body 时的兜底）
        """
        data = kwargs.get('data')
        if isinstance(data, dict) and csrf.FIELD_NAME not in data:
            data = dict(data)
            data[csrf.FIELD_NAME] = self._token()
            kwargs['data'] = data
            return kwargs

        payload = kwargs.get('json')
        if isinstance(payload, dict) and csrf.FIELD_NAME not in payload:
            payload = dict(payload)
            payload[csrf.FIELD_NAME] = self._token()
            kwargs['json'] = payload
            return kwargs

        if data is None and payload is None:
            headers = dict(kwargs.get('headers') or {})
            headers.setdefault('X-CSRF-Token', self._token())
            kwargs['headers'] = headers
        return kwargs

    def post(self, *args, **kwargs):
        return self._client.post(*args, **self._prepare(kwargs))

    def put(self, *args, **kwargs):
        return self._client.put(*args, **self._prepare(kwargs))

    def patch(self, *args, **kwargs):
        return self._client.patch(*args, **self._prepare(kwargs))

    def delete(self, *args, **kwargs):
        return self._client.delete(*args, **self._prepare(kwargs))

    def __getattr__(self, name):
        """其余方法（get / session_transaction / set_cookie ...）原样透传。"""
        return getattr(self._client, name)


def make_client(app):
    """创建一个会自动注入 CSRF 令牌的测试客户端。"""
    return CsrfClient(app.test_client())


# ============================================================
# 辅助函数
# ============================================================

def reset_database():
    """重置测试数据库：删除所有表并重新建表初始化。

    V2：evidence / blockers 两张新表也一并清空——tasks 表 DROP 后自增 ID
    从 1 重新开始，若不清空这两张表会残留指向旧 task_id 的脏数据。
    """
    # email_queue / email_log 一并清空：这两张表跨测试残留会让
    # 「队列中待发 = 0」「发送记录为空」这类断言随机失败。
    conn = db.get_db()
    conn.execute('PRAGMA foreign_keys = OFF')
    for table in ['messages', 'progress_logs', 'evidence', 'blockers',
                  'tasks', 'users', 'system_config',
                  'email_queue', 'email_log']:
        conn.execute('DROP TABLE IF EXISTS ' + table)
    conn.commit()
    conn.execute('PRAGMA foreign_keys = ON')
    models.init_db()


def create_test_user(username='testuser', display_name='测试用户',
                     password='test123456', role='owner'):
    """创建测试用户，返回 user_id。"""
    password_hash = auth.hash_password(password)
    return models.create_user(username, display_name, password_hash, role=role)


def make_future_date(days=7):
    """生成未来 N 天的日期字符串 YYYY-MM-DD。"""
    return (datetime.now() + timedelta(days=days)).strftime('%Y-%m-%d')


def make_past_date(days=1):
    """生成过去 N 天的日期字符串 YYYY-MM-DD。"""
    return (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')


def create_task_direct(title, created_by, assignee, priority='medium',
                       due_date=None, status='pending', created_days_ago=0):
    """直接在数据库创建任务（可自定义创建时间），返回 task_id。"""
    if due_date is None:
        due_date = make_future_date(7)
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    created_at = (datetime.now() - timedelta(days=created_days_ago)).strftime('%Y-%m-%d %H:%M:%S')
    task_id = db.execute(
        "INSERT INTO tasks (title, description, created_by, assignee, status, priority, "
        "due_date, created_at, updated_at, is_overdue) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0)",
        (title, '测试描述', created_by, assignee, status, priority, due_date, created_at, now)
    )
    return task_id


# ============================================================
# 1. 数据库初始化测试
# ============================================================

class TestDatabaseInit(unittest.TestCase):
    """测试数据库初始化：建表、索引、默认配置。"""

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()

    def test_five_tables_created(self):
        """验证 5 张表全部创建。"""
        tables = db.query(
            "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
        )
        table_names = [t['name'] for t in tables]
        expected = {'users', 'tasks', 'progress_logs', 'messages', 'system_config'}
        self.assertTrue(expected.issubset(set(table_names)),
                        '缺少表，期望 %s，实际 %s' % (expected, set(table_names)))

    def test_indexes_created(self):
        """验证索引全部创建。"""
        indexes = db.query(
            "SELECT name FROM sqlite_master WHERE type='index' AND name LIKE 'idx_%' ORDER BY name"
        )
        index_names = set(i['name'] for i in indexes)
        expected_indexes = {
            'idx_tasks_status', 'idx_tasks_assignee', 'idx_tasks_due_date',
            'idx_tasks_is_overdue', 'idx_tasks_created_by',
            'idx_logs_task_id', 'idx_logs_operated_at',
            'idx_messages_recipient', 'idx_messages_is_read', 'idx_messages_created',
        }
        missing = expected_indexes - index_names
        self.assertEqual(missing, set(), '缺少索引: %s' % missing)

    def test_default_config_inserted(self):
        """验证默认配置数据写入。"""
        due_days = models.get_config('warning_due_days')
        inactive_days = models.get_config('warning_inactive_days')
        scan_interval = models.get_config('scan_interval_seconds')
        scan_time = models.get_config('warning_scan_time')

        self.assertEqual(due_days, '3', '到期预警天数默认应为 3')
        self.assertEqual(inactive_days, '7', '待激活预警天数默认应为 7')
        self.assertEqual(scan_interval, '300', '扫描间隔默认应为 300')
        self.assertEqual(scan_time, '09:00', '扫描时间默认应为 09:00')

    def test_init_db_idempotent(self):
        """验证重复调用 init_db 不会破坏数据。"""
        # 先插入一条数据
        create_test_user('user1', '用户1')
        count_before = models.count_users()
        self.assertEqual(count_before, 1)

        # 再次初始化
        models.init_db()

        # 数据应仍然存在
        count_after = models.count_users()
        self.assertEqual(count_after, 1, '重复初始化不应破坏已有数据')


# ============================================================
# 2. 状态机转换测试
# ============================================================

class TestStateMachine(unittest.TestCase):
    """测试状态机：转换合法性、权限校验、副作用。"""

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()
        self.admin_id = create_test_user('admin', '管理员', role='admin')
        self.owner_id = create_test_user('owner', '负责人', role='owner')

    # --- 合法转换 ---

    def test_pending_to_in_progress(self):
        """pending → in_progress 合法。"""
        ok, _ = validate_transition('pending', 'in_progress', 'owner')
        self.assertTrue(ok)

    def test_pending_to_closed(self):
        """pending → closed 合法。"""
        ok, _ = validate_transition('pending', 'closed', 'owner')
        self.assertTrue(ok)

    def test_pending_to_cancelled(self):
        """pending → cancelled 合法。"""
        ok, _ = validate_transition('pending', 'cancelled', 'owner')
        self.assertTrue(ok)

    def test_in_progress_to_closed(self):
        """in_progress → closed 合法。"""
        ok, _ = validate_transition('in_progress', 'closed', 'owner')
        self.assertTrue(ok)

    def test_in_progress_to_cancelled(self):
        """in_progress → cancelled 合法。"""
        ok, _ = validate_transition('in_progress', 'cancelled', 'owner')
        self.assertTrue(ok)

    def test_overdue_to_in_progress(self):
        """overdue → in_progress 合法。"""
        ok, _ = validate_transition('overdue', 'in_progress', 'owner')
        self.assertTrue(ok)

    def test_overdue_to_closed(self):
        """overdue → closed 合法。"""
        ok, _ = validate_transition('overdue', 'closed', 'owner')
        self.assertTrue(ok)

    def test_overdue_to_cancelled(self):
        """overdue → cancelled 合法。"""
        ok, _ = validate_transition('overdue', 'cancelled', 'owner')
        self.assertTrue(ok)

    def test_closed_to_in_progress_admin(self):
        """closed → in_progress 合法（仅 admin）。"""
        ok, _ = validate_transition('closed', 'in_progress', 'admin')
        self.assertTrue(ok)

    def test_cancelled_to_pending_admin(self):
        """cancelled → pending 合法（仅 admin）。"""
        ok, _ = validate_transition('cancelled', 'pending', 'admin')
        self.assertTrue(ok)

    # --- 非法转换 ---

    def test_pending_to_overdue_manual(self):
        """pending → overdue 不允许手动转换（应由系统自动触发）。"""
        ok, reason = validate_transition('pending', 'overdue', 'owner')
        self.assertFalse(ok, 'pending→overdue 不应允许手动转换')

    def test_in_progress_to_pending(self):
        """in_progress → pending 不允许。"""
        ok, reason = validate_transition('in_progress', 'pending', 'owner')
        self.assertFalse(ok)

    def test_closed_to_cancelled(self):
        """closed → cancelled 不允许。"""
        ok, reason = validate_transition('closed', 'cancelled', 'admin')
        self.assertFalse(ok)

    def test_same_status(self):
        """相同状态不允许转换。"""
        ok, reason = validate_transition('pending', 'pending', 'owner')
        self.assertFalse(ok)

    # --- 权限校验 ---

    def test_owner_cannot_reopen_closed(self):
        """owner 尝试 closed → in_progress 应被拒绝（权限不足）。"""
        ok, reason = validate_transition('closed', 'in_progress', 'owner')
        self.assertFalse(ok)
        self.assertIn('管理员', reason)

    def test_owner_cannot_reactivate_cancelled(self):
        """owner 尝试 cancelled → pending 应被拒绝（权限不足）。"""
        ok, reason = validate_transition('cancelled', 'pending', 'owner')
        self.assertFalse(ok)
        self.assertIn('管理员', reason)

    # --- get_allowed_transitions ---

    def test_get_allowed_transitions_owner(self):
        """owner 的可转换列表不应包含管理员专属转换。"""
        # closed 状态下，owner 不应看到任何可转换状态
        allowed = get_allowed_transitions('closed', 'owner')
        self.assertEqual(allowed, [], 'owner 在 closed 状态下不应有可转换状态')

    def test_get_allowed_transitions_admin_closed(self):
        """admin 在 closed 状态下应能看到 in_progress。"""
        allowed = get_allowed_transitions('closed', 'admin')
        statuses = [a['status'] for a in allowed]
        self.assertIn('in_progress', statuses)

    def test_get_allowed_transitions_pending(self):
        """pending 状态下应有 3 个可转换状态。"""
        allowed = get_allowed_transitions('pending', 'owner')
        statuses = set(a['status'] for a in allowed)
        self.assertEqual(statuses, {'in_progress', 'closed', 'cancelled'})

    # --- change_task_status 副作用 ---

    def test_change_status_writes_progress_log(self):
        """状态变更应写入进度日志。"""
        task_id = create_task_direct('测试任务', self.admin_id, self.owner_id)
        ok, _ = change_task_status(task_id, 'in_progress', self.owner_id, 'owner', '开始执行')
        self.assertTrue(ok)

        logs = models.get_progress_logs(task_id)
        self.assertEqual(len(logs), 1)
        self.assertEqual(logs[0]['status_from'], 'pending')
        self.assertEqual(logs[0]['status_to'], 'in_progress')
        self.assertEqual(logs[0]['progress_note'], '开始执行')

    def test_change_status_sets_closed_at(self):
        """闭环时应记录 closed_at 时间。"""
        task_id = create_task_direct('闭环测试', self.admin_id, self.owner_id)
        ok, _ = change_task_status(task_id, 'closed', self.owner_id, 'owner', '已完成')
        self.assertTrue(ok)

        task = models.get_task(task_id)
        self.assertIsNotNone(task['closed_at'], '闭环后 closed_at 不应为空')

    def test_change_status_clears_closed_at_on_reopen(self):
        """重新打开时应清空 closed_at。"""
        task_id = create_task_direct('重开测试', self.admin_id, self.owner_id)
        # 先闭环
        change_task_status(task_id, 'closed', self.owner_id, 'owner', '完成')
        task = models.get_task(task_id)
        self.assertIsNotNone(task['closed_at'])

        # admin 重新打开
        ok, _ = change_task_status(task_id, 'in_progress', self.admin_id, 'admin', '需返工')
        self.assertTrue(ok)

        task = models.get_task(task_id)
        self.assertIsNone(task['closed_at'], '重新打开后 closed_at 应清空')
        self.assertEqual(task['status'], 'in_progress')

    def test_change_status_creates_message(self):
        """状态变更应生成通知消息给相关方。"""
        task_id = create_task_direct('消息测试', self.admin_id, self.owner_id)
        # owner 变更状态 → 应通知 admin（创建人）
        ok, _ = change_task_status(task_id, 'in_progress', self.owner_id, 'owner', '开始')
        self.assertTrue(ok)

        messages = models.get_messages(self.admin_id)
        self.assertEqual(len(messages), 1)
        self.assertEqual(messages[0]['type'], 'status_change')
        self.assertIn('进行中', messages[0]['content'])

    def test_change_status_no_self_message(self):
        """自己创建自己负责的任务，状态变更不给自己发消息。"""
        task_id = create_task_direct('自建任务', self.owner_id, self.owner_id)
        ok, _ = change_task_status(task_id, 'in_progress', self.owner_id, 'owner', '开始')
        self.assertTrue(ok)

        # owner 不应收到消息（创建人和负责人都是自己）
        messages = models.get_messages(self.owner_id)
        self.assertEqual(len(messages), 0, '不应给自己发消息')

    def test_change_status_overdue_clears_flag(self):
        """从逾期恢复时应清除 is_overdue 标记。"""
        task_id = create_task_direct('逾期恢复', self.admin_id, self.owner_id, status='overdue')
        # 手动设置 is_overdue
        models.update_task(task_id, is_overdue=1)

        ok, _ = change_task_status(task_id, 'in_progress', self.owner_id, 'owner', '恢复推进')
        self.assertTrue(ok)

        task = models.get_task(task_id)
        self.assertEqual(task['is_overdue'], 0, '恢复后 is_overdue 应为 0')

    def test_change_status_nonexistent_task(self):
        """变更不存在的任务应返回失败。"""
        ok, reason = change_task_status(99999, 'in_progress', self.owner_id, 'owner')
        self.assertFalse(ok)
        self.assertIn('不存在', reason)


# ============================================================
# 3. 应用启动与路由测试
# ============================================================

class TestAppStartup(unittest.TestCase):
    """测试 Flask 应用创建、蓝图注册、路由可访问性。"""

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()
        self.client = make_client(self.app)

    def test_app_created(self):
        """Flask 应用应成功创建。"""
        self.assertIsNotNone(self.app)

    def test_blueprints_registered(self):
        """所有 7 个蓝图应注册。"""
        blueprint_names = set(self.app.blueprints.keys())
        expected = {'auth', 'task', 'progress', 'message', 'user', 'settings', 'dashboard'}
        self.assertTrue(expected.issubset(blueprint_names),
                        '缺少蓝图，期望 %s，实际 %s' % (expected, blueprint_names))

    def test_root_redirect_to_setup(self):
        """无管理员时根路由应重定向到 setup。"""
        resp = self.client.get('/')
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/setup', resp.headers.get('Location', ''))

    def test_root_redirect_to_login(self):
        """有管理员时根路由应重定向到 login（未登录）。"""
        create_test_user('admin', '管理员', role='admin')
        resp = self.client.get('/')
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp.headers.get('Location', ''))

    def test_login_page_accessible(self):
        """登录页应可访问。"""
        create_test_user('admin', '管理员', role='admin')
        resp = self.client.get('/login')
        self.assertEqual(resp.status_code, 200)

    def test_setup_page_accessible(self):
        """无管理员时 setup 页应可访问。"""
        resp = self.client.get('/setup')
        self.assertEqual(resp.status_code, 200)

    def test_protected_route_redirects(self):
        """未登录访问受保护路由应重定向到 login。"""
        resp = self.client.get('/tasks')
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp.headers.get('Location', ''))

    def test_dashboard_protected(self):
        """未登录访问 dashboard 应重定向到 login。"""
        resp = self.client.get('/dashboard')
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp.headers.get('Location', ''))

    def test_error_handlers_registered(self):
        """错误处理器应注册（403/404/500）。"""
        # 404 处理器
        resp = self.client.get('/nonexistent-page')
        self.assertEqual(resp.status_code, 404)


# ============================================================
# 4. 端到端功能测试
# ============================================================

class TestEndToEnd(unittest.TestCase):
    """端到端功能测试：模拟完整业务流程。"""

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()
        self.client = make_client(self.app)

    def _create_admin_via_setup(self):
        """通过初始化向导创建管理员。"""
        return self.client.post('/setup', data={
            'username': 'admin',
            'display_name': '管理员',
            'password': 'admin123456',
            'confirm_password': 'admin123456',
        }, follow_redirects=False)

    def _login(self, username='admin', password='admin123456'):
        """登录。"""
        return self.client.post('/login', data={
            'username': username,
            'password': password,
        }, follow_redirects=False)

    def _login_as(self, username, password):
        """以指定账号登录，返回响应对象。"""
        return self.client.post('/login', data={
            'username': username,
            'password': password,
        }, follow_redirects=False)

    def test_full_flow(self):
        """完整流程：setup → login → 创建任务 → 列表 → 详情 → 状态变更 →
        进度 → 消息 → 标记已读 → 发消息 → 个人设置 → 系统设置 →
        用户管理 → 仪表盘 → 登出。"""

        # 1. 访问首页 → 重定向到 setup
        resp = self.client.get('/')
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/setup', resp.headers.get('Location', ''))

        # 2. 创建管理员
        resp = self._create_admin_via_setup()
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp.headers.get('Location', ''))
        self.assertTrue(models.has_admin())

        # 3. 登录
        resp = self._login()
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/dashboard', resp.headers.get('Location', ''))

        # 4. 访问仪表盘
        resp = self.client.get('/dashboard')
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b'\xe6\xa6\x82\xe8\xa7\x88', resp.data)  # "概览"

        # 5. 访问任务列表
        resp = self.client.get('/tasks')
        self.assertEqual(resp.status_code, 200)

        # 6. 创建任务
        resp = self.client.post('/tasks/new', data={
            'title': '督办系统测试任务',
            'description': '这是一个测试任务',
            'assignee': 1,
            'priority': 'high',
            'due_date': make_future_date(7),
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)
        # 创建任务后跳转到任务列表页（避免直接访问 detail 模板的丑陋界面）
        self.assertIn('/tasks', resp.headers.get('Location', ''))
        self.assertNotIn('/tasks/1', resp.headers.get('Location', ''))

        # 7. 查看任务详情
        resp = self.client.get('/tasks/1')
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b'\xe7\x9d\xa3\xe5\x8a\x9e', resp.data)  # 包含"督办"

        # 8. 变更任务状态 → in_progress
        resp = self.client.post('/tasks/1/status', data={
            'status': 'in_progress',
            'progress_note': '已开始执行',
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)
        task = models.get_task(1)
        self.assertEqual(task['status'], 'in_progress')

        # 9. 提交进度备注
        resp = self.client.post('/tasks/1/progress', data={
            'progress_note': '进度更新：完成50%',
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)
        logs = models.get_progress_logs(1)
        self.assertGreaterEqual(len(logs), 2)  # 状态变更日志 + 进度日志

        # 10. 查看消息中心（V2：消息中心页由顶栏铃铛抽屉替代，旧链接 301 永久重定向到仪表盘）
        resp = self.client.get('/messages', follow_redirects=False)
        self.assertEqual(resp.status_code, 301)
        self.assertIn('/dashboard', resp.headers.get('Location', ''))

        # 11. 标记消息已读（管理员自己创建任务指派给自己，可能有消息）
        # 先获取消息
        messages = models.get_messages(1)
        if messages:
            msg_id = messages[0]['message_id']
            resp = self.client.post('/messages/%d/read' % msg_id)
            self.assertEqual(resp.status_code, 200)
            data = resp.get_json()
            self.assertTrue(data['success'])

        # 12. 管理员发消息（需要先创建一个接收人）
        create_test_user('owner1', '负责人甲', password='owner123456', role='owner')
        resp = self.client.post('/messages/send', data={
            'recipient': 2,
            'content': '请尽快处理测试任务',
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)
        # 验证消息已发送
        owner_msgs = models.get_messages(2)
        self.assertGreater(len(owner_msgs), 0)

        # 13. 修改个人设置
        resp = self.client.post('/settings/profile', data={
            'display_name': '超级管理员',
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)
        user = models.get_user(1)
        self.assertEqual(user['display_name'], '超级管理员')

        # 14. 修改系统设置（预警天数）
        resp = self.client.post('/settings/system', data={
            'warning_due_days': '5',
            'warning_inactive_days': '14',
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(models.get_config('warning_due_days'), '5')
        self.assertEqual(models.get_config('warning_inactive_days'), '14')

        # 15. 用户管理 - 新增用户
        resp = self.client.post('/users/new', data={
            'username': 'owner2',
            'display_name': '负责人乙',
            'password': 'owner123456',
            'role': 'owner',
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(models.count_users(), 3)

        # 16. 查看用户管理页
        resp = self.client.get('/users')
        self.assertEqual(resp.status_code, 200)

        # 17. 登出
        resp = self.client.get('/logout')
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp.headers.get('Location', ''))

        # 18. 登出后访问受保护页面 → 重定向到 login
        resp = self.client.get('/tasks')
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp.headers.get('Location', ''))

    def test_login_wrong_password(self):
        """错误密码登录应失败（V2：返回 401 状态码 + 页面通用错误提示）。"""
        create_test_user('admin', '管理员', password='admin123456', role='admin')
        resp = self.client.post('/login', data={
            'username': 'admin',
            'password': 'wrongpassword',
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 401)  # V2：登录失败返回 401
        self.assertIn(b'\xe9\x94\x99\xe8\xaf\xaf', resp.data)  # 包含"错误"

    def test_login_disabled_user(self):
        """停用的用户不能登录（V2：返回 401 状态码 + 页面通用错误提示）。"""
        uid = create_test_user('admin', '管理员', password='admin123456', role='admin')
        # 停用用户
        models.toggle_user_active(uid)

        resp = self.client.post('/login', data={
            'username': 'admin',
            'password': 'admin123456',
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 401)  # V2：登录失败返回 401
        self.assertIn(b'\xe9\x94\x99\xe8\xaf\xaf', resp.data)  # 包含"错误"

    def test_create_task_owner_assigns_to_self(self):
        """owner 创建任务时只能指派给自己。"""
        admin_id = create_test_user('admin', '管理员', password='admin123456', role='admin')
        owner_id = create_test_user('owner', '负责人', password='owner123456', role='owner')

        self._login_as('owner', 'owner123456')

        # owner 尝试指派给 admin（应被强制指派给自己）
        resp = self.client.post('/tasks/new', data={
            'title': 'owner自建任务',
            'description': '',
            'assignee': admin_id,  # 尝试指派给 admin
            'priority': 'medium',
            'due_date': make_future_date(7),
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)

        # 验证任务负责人是 owner 自己
        task = models.get_task(1)
        self.assertEqual(task['assignee'], owner_id,
                         'owner 创建的任务负责人应被强制设为自己')

    def test_task_assignment_creates_message(self):
        """管理员指派任务给负责人应生成指派消息。"""
        admin_id = create_test_user('admin', '管理员', password='admin123456', role='admin')
        owner_id = create_test_user('owner', '负责人', password='owner123456', role='owner')

        self._login_as('admin', 'admin123456')

        self.client.post('/tasks/new', data={
            'title': '指派任务',
            'description': '',
            'assignee': owner_id,
            'priority': 'high',
            'due_date': make_future_date(7),
        }, follow_redirects=False)

        # 验证 owner 收到指派消息
        owner_msgs = models.get_messages(owner_id)
        self.assertGreater(len(owner_msgs), 0)
        self.assertEqual(owner_msgs[0]['type'], 'assignment')

    def test_mark_all_read(self):
        """全部标记已读。"""
        admin_id = create_test_user('admin', '管理员', password='admin123456', role='admin')
        owner_id = create_test_user('owner', '负责人', password='owner123456', role='owner')

        self._login_as('admin', 'admin123456')
        # 创建任务指派给 owner → owner 收到消息
        self.client.post('/tasks/new', data={
            'title': '任务1', 'assignee': owner_id,
            'priority': 'medium', 'due_date': make_future_date(7),
        }, follow_redirects=False)
        self.client.post('/tasks/new', data={
            'title': '任务2', 'assignee': owner_id,
            'priority': 'medium', 'due_date': make_future_date(7),
        }, follow_redirects=False)

        # 以 owner 登录
        self.client.get('/logout')
        self._login_as('owner', 'owner123456')

        # 全部标记已读
        resp = self.client.post('/messages/read-all')
        self.assertEqual(resp.status_code, 200)
        data = resp.get_json()
        self.assertTrue(data['success'])
        self.assertGreaterEqual(data['count'], 2)

        # 验证未读数为 0
        self.assertEqual(models.get_unread_count(owner_id), 0)

    def test_unread_count_api(self):
        """未读消息数 API。"""
        admin_id = create_test_user('admin', '管理员', password='admin123456', role='admin')
        owner_id = create_test_user('owner', '负责人', password='owner123456', role='owner')

        self._login_as('admin', 'admin123456')
        self.client.post('/tasks/new', data={
            'title': '任务', 'assignee': owner_id,
            'priority': 'medium', 'due_date': make_future_date(7),
        }, follow_redirects=False)

        self.client.get('/logout')
        self._login_as('owner', 'owner123456')

        resp = self.client.get('/messages/unread-count')
        self.assertEqual(resp.status_code, 200)
        data = resp.get_json()
        self.assertGreaterEqual(data['count'], 1)

    def test_user_toggle(self):
        """管理员停用/启用用户。"""
        create_test_user('admin', '管理员', password='admin123456', role='admin')
        owner_id = create_test_user('owner', '负责人', password='owner123456', role='owner')

        self._login_as('admin', 'admin123456')
        resp = self.client.post('/users/%d/toggle' % owner_id, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)

        user = models.get_user(owner_id)
        self.assertEqual(user['is_active'], 0, '用户应被停用')

    def test_user_reset_password(self):
        """管理员重置用户密码。"""
        create_test_user('admin', '管理员', password='admin123456', role='admin')
        owner_id = create_test_user('owner', '负责人', password='owner123456', role='owner')

        self._login_as('admin', 'admin123456')
        resp = self.client.post('/users/%d/reset-password' % owner_id, data={
            'new_password': 'newpass123456',
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)

        # 验证新密码可登录
        self.client.get('/logout')
        resp = self._login_as('owner', 'newpass123456')
        self.assertEqual(resp.status_code, 302)  # 登录成功

    def test_progress_logs_json(self):
        """进度记录 JSON 接口。"""
        admin_id = create_test_user('admin', '管理员', password='admin123456', role='admin')
        self._login_as('admin', 'admin123456')

        self.client.post('/tasks/new', data={
            'title': '日志测试', 'assignee': admin_id,
            'priority': 'medium', 'due_date': make_future_date(7),
        }, follow_redirects=False)

        # 变更状态
        self.client.post('/tasks/1/status', data={
            'status': 'in_progress', 'progress_note': '开始',
        }, follow_redirects=False)

        # 获取日志 JSON
        resp = self.client.get('/tasks/1/logs')
        self.assertEqual(resp.status_code, 200)
        data = resp.get_json()
        self.assertTrue(data['success'])
        self.assertGreaterEqual(len(data['data']), 1)

    def test_task_delete_cancelled_only(self):
        """仅允许删除已撤销的任务。"""
        admin_id = create_test_user('admin', '管理员', password='admin123456', role='admin')
        self._login_as('admin', 'admin123456')

        # 创建一个 pending 任务
        self.client.post('/tasks/new', data={
            'title': '待删除', 'assignee': admin_id,
            'priority': 'medium', 'due_date': make_future_date(7),
        }, follow_redirects=False)

        # 尝试删除 pending 任务（应被拒绝）
        resp = self.client.post('/tasks/1/delete', follow_redirects=False)
        self.assertEqual(resp.status_code, 302)  # 重定向但不删除
        self.assertIsNotNone(models.get_task(1), 'pending 任务不应被删除')

        # 撤销任务
        self.client.post('/tasks/1/status', data={
            'status': 'cancelled', 'progress_note': '取消',
        }, follow_redirects=False)

        # 删除已撤销任务
        resp = self.client.post('/tasks/1/delete', follow_redirects=False)
        self.assertEqual(resp.status_code, 302)
        self.assertIsNone(models.get_task(1), '已撤销任务应被删除')

    def test_edit_task(self):
        """编辑任务字段。"""
        admin_id = create_test_user('admin', '管理员', password='admin123456', role='admin')
        self._login_as('admin', 'admin123456')

        self.client.post('/tasks/new', data={
            'title': '原标题', 'assignee': admin_id,
            'priority': 'medium', 'due_date': make_future_date(7),
        }, follow_redirects=False)

        resp = self.client.post('/tasks/1/edit', data={
            'title': '新标题',
            'description': '新描述',
            'assignee': admin_id,
            'priority': 'high',
            'due_date': make_future_date(14),
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)

        task = models.get_task(1)
        self.assertEqual(task['title'], '新标题')
        self.assertEqual(task['priority'], 'high')


# ============================================================
# 5. 权限测试
# ============================================================

class TestPermissions(unittest.TestCase):
    """测试权限控制：未登录拦截、角色隔离、数据级权限。"""

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()
        self.client = make_client(self.app)
        self.admin_id = create_test_user('admin', '管理员', password='admin123456', role='admin')
        self.owner_id = create_test_user('owner', '负责人', password='owner123456', role='owner')

    def _login_as_admin(self):
        self.client.post('/login', data={
            'username': 'admin', 'password': 'admin123456',
        }, follow_redirects=False)

    def _login_as_owner(self):
        self.client.post('/login', data={
            'username': 'owner', 'password': 'owner123456',
        }, follow_redirects=False)

    def _login_as(self, username, password):
        """以指定账号登录。"""
        self.client.post('/login', data={
            'username': username, 'password': password,
        }, follow_redirects=False)

    def test_unauthenticated_tasks_redirect(self):
        """未登录访问 /tasks 重定向到 login。"""
        resp = self.client.get('/tasks')
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp.headers.get('Location', ''))

    def test_unauthenticated_dashboard_redirect(self):
        """未登录访问 /dashboard 重定向到 login。"""
        resp = self.client.get('/dashboard')
        self.assertEqual(resp.status_code, 302)

    def test_unauthenticated_messages_redirect(self):
        """未登录访问 /messages 重定向到 login。"""
        resp = self.client.get('/messages')
        self.assertEqual(resp.status_code, 302)

    def test_owner_access_users_403(self):
        """owner 访问 /users 应返回 403。"""
        self._login_as_owner()
        resp = self.client.get('/users')
        self.assertEqual(resp.status_code, 403)

    def test_owner_access_system_settings_403(self):
        """owner 访问 /settings/system 应返回 403。"""
        self._login_as_owner()
        resp = self.client.get('/settings/system')
        self.assertEqual(resp.status_code, 403)

    def test_owner_access_send_message_403(self):
        """owner 访问 /messages/send 应返回 403。"""
        self._login_as_owner()
        resp = self.client.get('/messages/send')
        self.assertEqual(resp.status_code, 403)

    def test_owner_cannot_edit_others_task(self):
        """owner 不能编辑 admin 创建的他人任务。"""
        self._login_as_admin()
        # admin 创建任务指派给 owner
        self.client.post('/tasks/new', data={
            'title': 'admin创建', 'assignee': self.owner_id,
            'priority': 'medium', 'due_date': make_future_date(7),
        }, follow_redirects=False)
        task_id = models.get_task(1)['task_id']

        # 切换到 owner
        self.client.get('/logout')
        self._login_as_owner()

        # owner 尝试编辑 admin 创建的任务（owner 是负责人，可以编辑自己的）
        # 但如果任务指派给另一个 owner，则不能编辑
        create_test_user('owner2', '负责人乙', password='owner123456', role='owner')

        self.client.get('/logout')
        self._login_as('owner2', 'owner123456')

        # owner2 尝试编辑 owner1 的任务
        resp = self.client.post('/tasks/%d/edit' % task_id, data={
            'title': '篡改标题', 'assignee': 3,
            'priority': 'low', 'due_date': make_future_date(1),
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)  # 重定向（拒绝）

        # 验证标题未被修改
        task = models.get_task(task_id)
        self.assertEqual(task['title'], 'admin创建', 'owner2 不应能修改他人任务')

    def test_owner_cannot_delete_task(self):
        """owner 不能删除任务（403）。"""
        self._login_as_admin()
        self.client.post('/tasks/new', data={
            'title': '测试删除', 'assignee': self.admin_id,
            'priority': 'medium', 'due_date': make_future_date(7),
        }, follow_redirects=False)

        self.client.get('/logout')
        self._login_as_owner()

        resp = self.client.post('/tasks/1/delete', follow_redirects=False)
        self.assertEqual(resp.status_code, 403)

    def test_owner_cannot_change_status_of_others_task(self):
        """owner 不能变更他人任务的状态。"""
        # admin 创建任务指派给 owner2
        create_test_user('owner2', '负责人乙', password='owner123456', role='owner')
        self._login_as_admin()
        self.client.post('/tasks/new', data={
            'title': '他人任务', 'assignee': 3,  # owner2
            'priority': 'medium', 'due_date': make_future_date(7),
        }, follow_redirects=False)

        # owner1 登录，尝试变更 owner2 的任务状态
        self.client.get('/logout')
        self._login_as_owner()

        resp = self.client.post('/tasks/1/status', data={
            'status': 'in_progress', 'progress_note': '篡改',
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)  # 重定向（拒绝）

        task = models.get_task(1)
        self.assertEqual(task['status'], 'pending', 'owner 不应能变更他人任务状态')

    def test_admin_can_edit_any_task(self):
        """admin 可以编辑任何人的任务。"""
        self._login_as_admin()
        self.client.post('/tasks/new', data={
            'title': '原任务', 'assignee': self.owner_id,
            'priority': 'medium', 'due_date': make_future_date(7),
        }, follow_redirects=False)

        resp = self.client.post('/tasks/1/edit', data={
            'title': 'admin修改', 'assignee': self.owner_id,
            'priority': 'urgent', 'due_date': make_future_date(3),
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)

        task = models.get_task(1)
        self.assertEqual(task['title'], 'admin修改')
        self.assertEqual(task['priority'], 'urgent')

    def test_owner_can_view_all_tasks(self):
        """owner 可以查看所有任务列表（Q4-B）。"""
        self._login_as_admin()
        # admin 创建多个任务
        self.client.post('/tasks/new', data={
            'title': '任务A', 'assignee': self.admin_id,
            'priority': 'medium', 'due_date': make_future_date(7),
        }, follow_redirects=False)
        self.client.post('/tasks/new', data={
            'title': '任务B', 'assignee': self.owner_id,
            'priority': 'medium', 'due_date': make_future_date(7),
        }, follow_redirects=False)

        self.client.get('/logout')
        self._login_as_owner()

        resp = self.client.get('/tasks')
        self.assertEqual(resp.status_code, 200)
        # owner 应能看到所有任务（包括 admin 的）
        self.assertIn(b'\xe4\xbb\xbb\xe5\x8a\xa1A', resp.data)  # "任务A"


# ============================================================
# 6. 分页测试（C1：每页 20 条）
# ============================================================

class TestPagination(unittest.TestCase):
    """测试分页功能（Web 端默认每页 10 条，由 config.TASKS_PER_PAGE 控制）。"""

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()
        self.client = make_client(self.app)
        self.admin_id = create_test_user('admin', '管理员', password='admin123456', role='admin')
        self.client.post('/login', data={
            'username': 'admin', 'password': 'admin123456',
        }, follow_redirects=False)

    def test_pagination_25_tasks(self):
        """创建 25 条任务，验证分页（第1页20条，第2页5条）。"""
        # 创建 25 条任务
        for i in range(25):
            self.client.post('/tasks/new', data={
                'title': '分页测试任务-%02d' % (i + 1),
                'assignee': self.admin_id,
                'priority': 'medium',
                'due_date': make_future_date(7 + i),
            }, follow_redirects=False)

        # 第 1 页应有 20 条
        tasks_page1, total = models.get_tasks(page=1, per_page=20)
        self.assertEqual(total, 25, '总任务数应为 25')
        self.assertEqual(len(tasks_page1), 20, '第1页应有 20 条任务')

        # 第 2 页应有 5 条
        tasks_page2, _ = models.get_tasks(page=2, per_page=20)
        self.assertEqual(len(tasks_page2), 5, '第2页应有 5 条任务')

        # 第 3 页应为空
        tasks_page3, _ = models.get_tasks(page=3, per_page=20)
        self.assertEqual(len(tasks_page3), 0, '第3页应为空')

    def test_pagination_via_web(self):
        """通过 Web 界面验证分页（每页 10 条，由 config.TASKS_PER_PAGE 控制）。"""
        for i in range(25):
            self.client.post('/tasks/new', data={
                'title': 'Web分页-%02d' % (i + 1),
                'assignee': self.admin_id,
                'priority': 'medium',
                'due_date': make_future_date(7 + i),
            }, follow_redirects=False)

        # 第 1 页（每页 10 条）：含 01~10
        resp = self.client.get('/tasks?page=1')
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b'Web\xe5\x88\x86\xe9\xa1\xb5-01', resp.data)  # "Web分页-01"
        self.assertIn(b'Web\xe5\x88\x86\xe9\xa1\xb5-10', resp.data)  # "Web分页-10"

        # 第 2 页：含 11~20，且不含第 1 页数据
        resp = self.client.get('/tasks?page=2')
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b'Web\xe5\x88\x86\xe9\xa1\xb5-11', resp.data)  # "Web分页-11"
        self.assertIn(b'Web\xe5\x88\x86\xe9\xa1\xb5-20', resp.data)  # "Web分页-20"
        self.assertNotIn(b'Web\xe5\x88\x86\xe9\xa1\xb5-01', resp.data)

        # 第 3 页：含 21~25
        resp = self.client.get('/tasks?page=3')
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b'Web\xe5\x88\x86\xe9\xa1\xb5-21', resp.data)  # "Web分页-21"
        self.assertIn(b'Web\xe5\x88\x86\xe9\xa1\xb5-25', resp.data)  # "Web分页-25"

    def test_pagination_invalid_page(self):
        """无效页码应回退到第 1 页。"""
        for i in range(5):
            self.client.post('/tasks/new', data={
                'title': '测试-%02d' % (i + 1),
                'assignee': self.admin_id,
                'priority': 'medium',
                'due_date': make_future_date(7),
            }, follow_redirects=False)

        # 页码为 0 → 应回退到 1
        resp = self.client.get('/tasks?page=0')
        self.assertEqual(resp.status_code, 200)

        # 页码为负 → 应回退到 1
        resp = self.client.get('/tasks?page=-1')
        self.assertEqual(resp.status_code, 200)

        # 页码非数字 → 应回退到 1
        resp = self.client.get('/tasks?page=abc')
        self.assertEqual(resp.status_code, 200)


# ============================================================
# 7. 预警引擎测试
# ============================================================

class TestWarningEngine(unittest.TestCase):
    """测试三层预警引擎。"""

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()
        self.admin_id = create_test_user('admin', '管理员', role='admin')
        self.owner_id = create_test_user('owner', '负责人', role='owner')

    def test_due_warning(self):
        """第一层：即将到期预警。"""
        # 创建一个 2 天后到期的 pending 任务（在 3 天窗口内）
        task_id = create_task_direct(
            '即将到期任务', self.admin_id, self.owner_id,
            due_date=make_future_date(2), status='pending'
        )
        import warning_engine
        warning_engine.run_warning_scan()

        # owner 应收到到期预警消息
        msgs = models.get_messages(self.owner_id)
        due_msgs = [m for m in msgs if m['type'] == 'warning_due']
        self.assertGreater(len(due_msgs), 0, '应生成到期预警消息')

    def test_overdue_warning(self):
        """第二层：已逾期预警（通知负责人+管理员）。"""
        task_id = create_task_direct(
            '已逾期任务', self.admin_id, self.owner_id,
            due_date=make_past_date(1), status='overdue'
        )
        models.update_task(task_id, is_overdue=1)

        import warning_engine
        warning_engine.run_warning_scan()

        # owner 应收到逾期预警
        owner_msgs = models.get_messages(self.owner_id)
        overdue_msgs = [m for m in owner_msgs if m['type'] == 'warning_overdue']
        self.assertGreater(len(overdue_msgs), 0, 'owner 应收到逾期预警')

        # admin 也应收到逾期预警
        admin_msgs = models.get_messages(self.admin_id)
        admin_overdue = [m for m in admin_msgs if m['type'] == 'warning_overdue']
        self.assertGreater(len(admin_overdue), 0, 'admin 应收到逾期预警')

    def test_inactive_warning(self):
        """第三层：长期待激活预警（创建超过7天未启动）。"""
        # 创建一个 7 天前创建的 pending 任务
        task_id = create_task_direct(
            '长期未启动任务', self.admin_id, self.owner_id,
            due_date=make_future_date(7), status='pending', created_days_ago=7
        )

        import warning_engine
        warning_engine.run_warning_scan()

        # owner 应收到待激活预警
        owner_msgs = models.get_messages(self.owner_id)
        inactive_msgs = [m for m in owner_msgs if m['type'] == 'warning_inactive']
        self.assertGreater(len(inactive_msgs), 0, '应生成待激活预警消息')

    def test_warning_dedup(self):
        """预警去重：同一天不重复发送。"""
        task_id = create_task_direct(
            '去重测试', self.admin_id, self.owner_id,
            due_date=make_future_date(2), status='pending'
        )

        import warning_engine
        # 第一次扫描
        warning_engine.run_warning_scan()
        msgs_after_first = models.get_messages(self.owner_id)
        count1 = len(msgs_after_first)

        # 第二次扫描（同一天）
        warning_engine.run_warning_scan()
        msgs_after_second = models.get_messages(self.owner_id)
        count2 = len(msgs_after_second)

        self.assertEqual(count1, count2, '同一天不应重复发送预警消息')

    def test_no_warning_for_closed_task(self):
        """终态任务不触发预警。"""
        task_id = create_task_direct(
            '已闭环任务', self.admin_id, self.owner_id,
            due_date=make_future_date(2), status='closed'
        )

        import warning_engine
        warning_engine.run_warning_scan()

        msgs = models.get_messages(self.owner_id)
        warning_msgs = [m for m in msgs if m['type'].startswith('warning_')]
        self.assertEqual(len(warning_msgs), 0, '已闭环任务不应触发预警')

    def test_trigger_overdue_warning(self):
        """触发逾期预警（任务被自动标记为逾期时）。"""
        task = models.get_task(create_task_direct(
            '自动逾期', self.admin_id, self.owner_id,
            due_date=make_past_date(1), status='pending'
        ))
        # 将 Row 转为 dict 传给 trigger_overdue_warning
        import warning_engine
        task_dict = dict(task)
        warning_engine.trigger_overdue_warning(task_dict)

        # owner 和 admin 都应收到消息
        owner_msgs = models.get_messages(self.owner_id)
        admin_msgs = models.get_messages(self.admin_id)
        self.assertGreater(len(owner_msgs), 0)
        self.assertGreater(len(admin_msgs), 0)


# ============================================================
# 8. 仪表盘统计测试
# ============================================================

class TestDashboard(unittest.TestCase):
    """测试仪表盘统计数据。"""

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()
        self.admin_id = create_test_user('admin', '管理员', role='admin')
        self.owner_id = create_test_user('owner', '负责人', role='owner')

    def test_dashboard_stats_empty(self):
        """空数据库的仪表盘统计。"""
        stats = models.get_dashboard_stats(self.admin_id)
        self.assertEqual(stats['total'], 0)
        self.assertEqual(stats['pending'], 0)
        self.assertEqual(stats['overdue'], 0)

    def test_dashboard_stats_with_data(self):
        """有数据时的仪表盘统计。"""
        # 创建不同状态的任务
        create_task_direct('任务1', self.admin_id, self.owner_id, status='pending')
        create_task_direct('任务2', self.admin_id, self.owner_id, status='in_progress')
        create_task_direct('任务3', self.admin_id, self.owner_id, status='overdue')
        create_task_direct('任务4', self.admin_id, self.owner_id, status='closed')
        create_task_direct('任务5', self.admin_id, self.owner_id, status='cancelled')

        stats = models.get_dashboard_stats(self.owner_id)
        # 总数不包含 cancelled
        self.assertEqual(stats['total'], 4)
        self.assertEqual(stats['pending'], 1)
        self.assertEqual(stats['in_progress'], 1)
        self.assertEqual(stats['overdue'], 1)
        self.assertEqual(stats['closed'], 1)
        self.assertEqual(stats['cancelled'], 1)
        # 我的待办（owner 有 3 个非终态任务）
        self.assertEqual(stats['my_pending'], 3)


# ============================================================
# 批量操作 + CSV 导出测试（P1-003 / P1-004）
# ============================================================

class TestBatchAndExport(unittest.TestCase):
    """测试任务批量操作（P1-003）和 Excel 导出（P1-004）。"""

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()
        self.client = make_client(self.app)
        # 创建 admin + 2 个 owner
        self.admin_id = create_test_user('admin', '管理员', role='admin')
        self.owner1_id = create_test_user('owner1', '张三', role='owner')
        self.owner2_id = create_test_user('owner2', '李四', role='owner')
        # 创建若干任务（admin 创建并指派给 owner1/owner2）
        self.t1 = create_task_direct('任务A', self.admin_id, self.owner1_id, status='pending')
        self.t2 = create_task_direct('任务B', self.admin_id, self.owner1_id, status='pending')
        self.t3 = create_task_direct('任务C', self.admin_id, self.owner2_id, status='pending')
        self.t4 = create_task_direct('任务D', self.admin_id, self.owner2_id, status='in_progress')

    def _login(self, username='admin'):
        """以指定账号登录。"""
        pwd = {'admin': 'test123456', 'owner1': 'test123456', 'owner2': 'test123456'}
        return self.client.post('/login', data={
            'username': username, 'password': pwd[username]
        }, follow_redirects=False)

    # ---------- Excel 导出测试 ----------

    def test_export_xlsx_requires_login(self):
        """未登录访问 /tasks/export 应重定向到 login。"""
        resp = self.client.get('/tasks/export', follow_redirects=False)
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp.headers.get('Location', ''))

    def test_export_xlsx_returns_xlsx(self):
        """登录后导出应返回 xlsx 附件（每列自适应列宽），含表头和数据。"""
        self._login('admin')
        resp = self.client.get('/tasks/export', follow_redirects=False)
        self.assertEqual(resp.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resp.content_type
        )
        # Content-Disposition 应为附件下载（双 fallback：ASCII filename + UTF-8 filename*）
        cd = resp.headers.get('Content-Disposition', '')
        self.assertIn('attachment', cd)
        from datetime import datetime
        date_str = datetime.now().strftime('%Y-%m-%d')
        self.assertIn(f'tasks_{date_str}.xlsx', cd)
        # 中文名通过 RFC 5987 filename* URL 编码传递（督办任务列表 的百分号编码）
        self.assertIn('UTF-8\'\'', cd)
        self.assertIn('%E7%9D%A3%E5%8A%9E%E4%BB%BB%E5%8A%A1%E5%88%97%E8%A1%A8', cd)
        # 解析 xlsx 内容（openpyxl）
        import io as _io
        from openpyxl import load_workbook
        wb = load_workbook(_io.BytesIO(resp.data))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # 表头
        header = rows[0]
        self.assertIn('任务ID', header)
        self.assertIn('标题', header)
        self.assertIn('负责人', header)
        self.assertIn('是否逾期', header)
        # 数据行（任务A，负责人张三，状态待启动）
        data = [r for r in rows[1:] if r and r[1] == '任务A']
        self.assertTrue(data, '应含任务A数据行')
        self.assertEqual(data[0][2], '张三')   # 负责人 display_name
        self.assertEqual(data[0][5], '待启动')  # 状态中文
        # 每列已设置列宽（自适应）
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            dim = ws.column_dimensions[col_letter].width
            self.assertIsNotNone(dim, f'列 {col_letter} 应已设置列宽')
            self.assertGreater(dim, 0)

    def test_export_xlsx_with_filter(self):
        """带 status=in_progress 筛选只导出进行中的任务。"""
        self._login('admin')
        resp = self.client.get('/tasks/export?status=in_progress', follow_redirects=False)
        self.assertEqual(resp.status_code, 200)
        import io as _io
        from openpyxl import load_workbook
        wb = load_workbook(_io.BytesIO(resp.data))
        ws = wb.active
        titles = [r[1] for r in ws.iter_rows(min_row=2, values_only=True) if r]
        # 只含任务D（in_progress），不含任务A/B/C（pending）
        self.assertIn('任务D', titles)
        self.assertNotIn('任务A', titles)
        self.assertNotIn('任务B', titles)
        self.assertNotIn('任务C', titles)

    # ---------- 批量操作权限/校验测试 ----------

    def test_batch_requires_login(self):
        """未登录 POST /tasks/batch 应重定向到 login。"""
        resp = self.client.post('/tasks/batch', data={
            'task_ids': [self.t1], 'action': 'change_status', 'target_status': 'in_progress'
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp.headers.get('Location', ''))

    def test_batch_no_selection(self):
        """未选任务应 flash 错误并重定向。"""
        self._login('admin')
        resp = self.client.post('/tasks/batch', data={
            'action': 'change_status', 'target_status': 'in_progress'
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/tasks', resp.headers.get('Location', ''))

    def test_batch_invalid_action(self):
        """非法 action 应 flash 错误。"""
        self._login('admin')
        resp = self.client.post('/tasks/batch', data={
            'task_ids': [self.t1], 'action': 'unknown_action'
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)

    # ---------- 批量改状态测试 ----------

    def test_batch_change_status_admin(self):
        """admin 批量改状态：4 个 pending/in_progress 任务全部成功转为 closed。"""
        self._login('admin')
        resp = self.client.post('/tasks/batch', data={
            'task_ids': [self.t1, self.t2, self.t3, self.t4],
            'action': 'change_status',
            'target_status': 'cancelled'
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)
        # 验证状态已变更
        for tid in [self.t1, self.t2, self.t3, self.t4]:
            task = models.get_task(tid)
            self.assertEqual(task['status'], 'cancelled', '任务 %d 应已撤销' % tid)

    def test_batch_change_status_owner_only_own(self):
        """owner1 批量改状态：自己的任务成功，别人的失败。"""
        self._login('owner1')
        # owner1 批量改 t1/t2（自己的）+ t3（owner2 的）+ t4（owner2 的）
        resp = self.client.post('/tasks/batch', data={
            'task_ids': [self.t1, self.t2, self.t3, self.t4],
            'action': 'change_status',
            'target_status': 'in_progress'
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)
        # t1/t2 成功转 in_progress
        self.assertEqual(models.get_task(self.t1)['status'], 'in_progress')
        self.assertEqual(models.get_task(self.t2)['status'], 'in_progress')
        # t3/t4 无权操作，状态不变
        self.assertEqual(models.get_task(self.t3)['status'], 'pending')
        self.assertEqual(models.get_task(self.t4)['status'], 'in_progress')

    def test_batch_change_status_invalid_transition(self):
        """批量改到非法目标状态（如已闭环→已撤销不允许 via 矩阵）应失败但不中断整体。"""
        # 先把 t1 改为 closed
        from state_machine import change_task_status
        change_task_status(self.t1, 'in_progress', self.admin_id, 'admin', '')
        change_task_status(self.t1, 'closed', self.admin_id, 'admin', '完成')
        # closed→cancelled 不在转换矩阵中，应失败
        self._login('admin')
        resp = self.client.post('/tasks/batch', data={
            'task_ids': [self.t1, self.t2],
            'action': 'change_status',
            'target_status': 'cancelled'
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)
        # t1 仍 closed（转换失败），t2 转为 cancelled（pending→cancelled 允许）
        self.assertEqual(models.get_task(self.t1)['status'], 'closed')
        self.assertEqual(models.get_task(self.t2)['status'], 'cancelled')

    # ---------- 批量指派测试 ----------

    def test_batch_reassign_admin(self):
        """admin 批量指派：把 t1/t2 从 owner1 转给 owner2。"""
        self._login('admin')
        resp = self.client.post('/tasks/batch', data={
            'task_ids': [self.t1, self.t2],
            'action': 'reassign',
            'new_assignee': str(self.owner2_id)
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)
        # 验证负责人已变更
        self.assertEqual(models.get_task(self.t1)['assignee'], self.owner2_id)
        self.assertEqual(models.get_task(self.t2)['assignee'], self.owner2_id)
        # 应生成指派消息通知 owner2
        msgs = models.get_messages(self.owner2_id, {})
        assign_msgs = [m for m in msgs if m['type'] == 'assignment']
        self.assertGreaterEqual(len(assign_msgs), 2, '应生成至少 2 条指派消息')

    def test_batch_reassign_owner_forbidden(self):
        """owner 不能批量指派（被拒，flash 错误）。"""
        self._login('owner1')
        resp = self.client.post('/tasks/batch', data={
            'task_ids': [self.t1],
            'action': 'reassign',
            'new_assignee': str(self.owner2_id)
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)
        # 负责人未变（操作被拒）
        self.assertEqual(models.get_task(self.t1)['assignee'], self.owner1_id)

    def test_batch_reassign_invalid_assignee(self):
        """批量指派到不存在/停用的负责人应失败。"""
        self._login('admin')
        resp = self.client.post('/tasks/batch', data={
            'task_ids': [self.t1],
            'action': 'reassign',
            'new_assignee': '99999'  # 不存在的 user_id
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)
        # 负责人未变
        self.assertEqual(models.get_task(self.t1)['assignee'], self.owner1_id)

    # ---------- 页面集成测试 ----------

    def test_list_page_has_export_link(self):
        """/tasks 页面应包含导出 Excel 链接。"""
        self._login('admin')
        resp = self.client.get('/tasks', follow_redirects=False)
        self.assertEqual(resp.status_code, 200)
        self.assertIn('导出 Excel', resp.data.decode('utf-8'))
        self.assertIn('/tasks/export', resp.data.decode('utf-8'))

    def test_list_page_has_batch_form(self):
        """/tasks 页面应包含批量操作表单和复选框。"""
        self._login('admin')
        resp = self.client.get('/tasks', follow_redirects=False)
        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode('utf-8')
        self.assertIn('batch-form', html)
        self.assertIn('select-all', html)
        self.assertIn('task-checkbox', html)
        self.assertIn('批量改状态', html)
        # admin 应能看到批量指派选项
        self.assertIn('批量指派负责人', html)


# ============================================================
# 9. V2 仪表盘：闭环率口径 / 时间范围 / 闭环矩阵 / 今日焦点
# ============================================================

class TestDashboardV2(unittest.TestCase):
    """V2 仪表盘统计：闭环率口径（含撤销场景）、时间范围筛选、闭环矩阵聚合。"""

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()
        self.admin_id = create_test_user('admin', '管理员', role='admin')
        self.owner1_id = create_test_user('owner1', '张三', role='owner')
        self.owner2_id = create_test_user('owner2', '李四', role='owner')

    def test_closure_rate_excludes_cancelled(self):
        """闭环率口径：closed ÷ (总数 - 已撤销)，撤销任务不计入分母。"""
        create_task_direct('待办1', self.admin_id, self.owner1_id, status='pending')
        create_task_direct('待办2', self.admin_id, self.owner1_id, status='pending')
        create_task_direct('已闭环1', self.admin_id, self.owner1_id, status='closed')
        create_task_direct('已闭环2', self.admin_id, self.owner1_id, status='closed')
        create_task_direct('已撤销', self.admin_id, self.owner1_id, status='cancelled')

        stats = models.get_dashboard_stats_v2('all')
        # 任务总数 = 5 - 1（撤销）= 4
        self.assertEqual(stats['total'], 4, '任务总数应为 4（不含已撤销）')
        self.assertEqual(stats['closed'], 2)
        self.assertEqual(stats['cancelled'], 1)
        # 闭环率 = 2 / 4 = 0.5
        self.assertEqual(stats['closure_rate'], 0.5,
                         '闭环率应为 2/4=0.5，撤销任务不计入分母')

        # 边界：全部任务都已撤销 → 分母为 0，闭环率应为 None
        create_task_direct('撤销2', self.admin_id, self.owner1_id, status='cancelled')
        create_task_direct('撤销3', self.admin_id, self.owner1_id, status='cancelled')
        # 此时非撤销任务仍是 4 条，再单独验证全撤销场景
        reset_database()
        create_test_user('admin', '管理员', role='admin')
        self.owner_id = create_test_user('owner', '负责人', role='owner')
        create_task_direct('仅撤销1', self.admin_id, self.owner_id, status='cancelled')
        create_task_direct('仅撤销2', self.admin_id, self.owner_id, status='cancelled')
        stats2 = models.get_dashboard_stats_v2('all')
        self.assertEqual(stats2['total'], 0, '全部撤销时任务总数应为 0')
        self.assertIsNone(stats2['closure_rate'], '分母为 0 时闭环率应为 None')

    def test_stats_time_range_filter_and_fallback(self):
        """时间范围筛选：week 只统计本周创建的任务；非法 range 回落到 all。"""
        # 10 天前创建的旧任务（必然早于本周一）+ 今天创建的新任务
        create_task_direct('旧任务', self.admin_id, self.owner1_id,
                           status='pending', created_days_ago=10)
        create_task_direct('新任务', self.admin_id, self.owner1_id, status='pending')

        stats_all = models.get_dashboard_stats_v2('all')
        self.assertEqual(stats_all['total'], 2, 'all 范围应包含 2 条任务')

        stats_week = models.get_dashboard_stats_v2('week')
        self.assertEqual(stats_week['total'], 1, 'week 范围应只包含本周创建的 1 条任务')

        # 非法 range 键应回落到 all（不过滤）
        stats_bogus = models.get_dashboard_stats_v2('not_a_range')
        self.assertEqual(stats_bogus['total'], 2, '非法 range 应回落到 all 口径')

    def test_dashboard_page_time_range(self):
        """仪表盘页面 ?range= 参数：本周视图的焦点列表不包含本周前创建的任务。"""
        self.client = make_client(self.app)
        self.client.post('/login', data={
            'username': 'admin', 'password': 'test123456',
        }, follow_redirects=False)

        create_task_direct('本周前旧任务', self.admin_id, self.owner1_id,
                           status='pending', created_days_ago=10)
        create_task_direct('本周新任务', self.admin_id, self.owner1_id, status='pending')

        # 本周视图：只含新任务
        resp = self.client.get('/dashboard?range=week')
        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode('utf-8')
        self.assertIn('本周新任务', html)
        self.assertNotIn('本周前旧任务', html)

        # 全部视图：两条都在
        resp = self.client.get('/dashboard?range=all')
        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode('utf-8')
        self.assertIn('本周新任务', html)
        self.assertIn('本周前旧任务', html)

    def test_closure_matrix_aggregation(self):
        """闭环矩阵：按负责人聚合任务数/闭环数/闭环率，按任务数降序。"""
        # 张三：2 待启动 + 1 已闭环 → 任务数 3，闭环 1，闭环率 1/3
        create_task_direct('张三-待办1', self.admin_id, self.owner1_id, status='pending')
        create_task_direct('张三-待办2', self.admin_id, self.owner1_id, status='pending')
        create_task_direct('张三-闭环', self.admin_id, self.owner1_id, status='closed')
        # 李四：1 已闭环 + 1 已撤销 → 任务数 1，闭环 1，闭环率 1.0
        create_task_direct('李四-闭环', self.admin_id, self.owner2_id, status='closed')
        create_task_direct('李四-撤销', self.admin_id, self.owner2_id, status='cancelled')

        matrix, total, total_pages, page = models.get_closure_matrix('all')

        self.assertEqual(total, 2, '应有 2 位负责人')
        self.assertEqual(len(matrix), 2)
        # 任务数降序：张三（3）在前
        self.assertEqual(matrix[0]['display_name'], '张三')
        self.assertEqual(matrix[0]['total'], 3)
        self.assertEqual(matrix[0]['pending'], 2)
        self.assertEqual(matrix[0]['closed'], 1)
        self.assertAlmostEqual(matrix[0]['closure_rate'], 1 / 3, places=4)

        self.assertEqual(matrix[1]['display_name'], '李四')
        self.assertEqual(matrix[1]['total'], 1, '李四任务数应为 1（撤销不计入）')
        self.assertEqual(matrix[1]['closed'], 1)
        self.assertEqual(matrix[1]['cancelled'], 1)
        self.assertEqual(matrix[1]['closure_rate'], 1.0)

    def test_today_focus_ordering_and_mine_flag(self):
        """今日督办焦点：已逾期 > 进行中 > 待启动排序，is_mine 标记，终态任务排除。"""
        create_task_direct('待启动任务', self.admin_id, self.owner1_id, status='pending')
        create_task_direct('进行中任务', self.admin_id, self.owner1_id, status='in_progress')
        create_task_direct('已逾期任务', self.admin_id, self.owner1_id, status='overdue')
        create_task_direct('已闭环任务', self.admin_id, self.owner1_id, status='closed')
        create_task_direct('已撤销任务', self.admin_id, self.owner1_id, status='cancelled')
        create_task_direct('管理员的任务', self.admin_id, self.admin_id, status='pending')

        focus = models.get_today_focus('all', user_id=self.owner1_id)
        titles = [f['title'] for f in focus]
        # 只含非终态任务
        self.assertEqual(len(focus), 4, '焦点应只含待启动/进行中/已逾期 4 条')
        self.assertNotIn('已闭环任务', titles)
        self.assertNotIn('已撤销任务', titles)
        # 排序：已逾期 → 进行中 → 待启动（管理员的任务也是 pending）
        self.assertEqual(focus[0]['title'], '已逾期任务')
        self.assertEqual(focus[1]['title'], '进行中任务')
        # is_mine：owner1 的任务为 True，admin 的任务为 False
        by_title = {f['title']: f for f in focus}
        self.assertTrue(by_title['待启动任务']['is_mine'], '自己负责的任务应标记 is_mine')
        self.assertFalse(by_title['管理员的任务']['is_mine'], '他人任务不应标记 is_mine')


# ============================================================
# 10. V2 行内编辑：POST /tasks/<id>/field 权限矩阵与校验
# ============================================================

class TestInlineFieldEdit(unittest.TestCase):
    """V2 行内编辑接口：admin 改任意任务 / owner 只能改自己的 / owner 改他人 403。"""

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()
        self.client = make_client(self.app)
        self.admin_id = create_test_user('admin', '管理员', password='admin123456', role='admin')
        self.owner1_id = create_test_user('owner1', '张三', password='owner123456', role='owner')
        self.owner2_id = create_test_user('owner2', '李四', password='owner123456', role='owner')
        # admin 创建任务指派给 owner1
        self.task_id = create_task_direct('行内编辑测试任务', self.admin_id, self.owner1_id)

    def _login(self, username, password):
        return self.client.post('/login', data={
            'username': username, 'password': password,
        }, follow_redirects=False)

    def _post_field(self, task_id, field, value):
        return self.client.post(
            '/tasks/%d/field' % task_id,
            json={'field': field, 'value': value},
        )

    def test_admin_edit_any_task_field(self):
        """admin 可行内编辑任意任务（含负责人变更 + 通知新负责人）。"""
        self._login('admin', 'admin123456')

        # 改标题
        resp = self._post_field(self.task_id, 'title', '管理员改过的标题')
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.get_json()['success'])
        self.assertEqual(models.get_task(self.task_id)['title'], '管理员改过的标题')

        # admin 把负责人从 owner1 改为 owner2，应成功并通知 owner2
        resp = self._post_field(self.task_id, 'assignee', self.owner2_id)
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.get_json()['success'])
        self.assertEqual(models.get_task(self.task_id)['assignee'], self.owner2_id)
        msgs = models.get_messages(self.owner2_id)
        assign_msgs = [m for m in msgs if m['type'] == 'assignment']
        self.assertGreater(len(assign_msgs), 0, '负责人变更应通知新负责人')

    def test_owner_edit_own_task_fields(self):
        """owner 可编辑自己任务的 title/progress_percent/risk_note/collaborators。"""
        self._login('owner1', 'owner123456')

        resp = self._post_field(self.task_id, 'title', '负责人改过的标题')
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.get_json()['success'])

        resp = self._post_field(self.task_id, 'progress_percent', 60)
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.get_json()['success'])
        self.assertEqual(resp.get_json().get('progress'), 60)

        resp = self._post_field(self.task_id, 'risk_note', '存在外部依赖风险')
        self.assertEqual(resp.status_code, 200)

        resp = self._post_field(self.task_id, 'collaborators', '财务部、法务部')
        self.assertEqual(resp.status_code, 200)

        task = models.get_task(self.task_id)
        self.assertEqual(task['title'], '负责人改过的标题')
        self.assertEqual(task['progress_percent'], 60)
        self.assertEqual(task['risk_note'], '存在外部依赖风险')
        self.assertEqual(task['collaborators'], '财务部、法务部')

    def test_owner_edit_others_task_field_403(self):
        """owner 行内编辑他人任务应返回 403，字段不被修改。"""
        self._login('owner2', 'owner123456')
        resp = self._post_field(self.task_id, 'title', '越权篡改标题')
        self.assertEqual(resp.status_code, 403)
        self.assertFalse(resp.get_json()['success'])
        self.assertEqual(models.get_task(self.task_id)['title'],
                         '行内编辑测试任务', 'owner2 不应能改他人任务标题')

    def test_field_validation_and_not_found(self):
        """行内编辑校验：进度越界/空标题/非法字段 400，任务不存在 404。"""
        self._login('admin', 'admin123456')

        # 进度 150 → 400
        resp = self._post_field(self.task_id, 'progress_percent', 150)
        self.assertEqual(resp.status_code, 400)
        self.assertFalse(resp.get_json()['success'])
        self.assertEqual(models.get_task(self.task_id)['progress_percent'], 0)

        # 空标题 → 400
        resp = self._post_field(self.task_id, 'title', '   ')
        self.assertEqual(resp.status_code, 400)

        # 白名单外字段 → 400
        resp = self._post_field(self.task_id, 'status', 'closed')
        self.assertEqual(resp.status_code, 400)
        self.assertEqual(models.get_task(self.task_id)['status'], 'pending')

        # 任务不存在 → 404
        resp = self._post_field(99999, 'title', '幽灵任务')
        self.assertEqual(resp.status_code, 404)

    def test_owner_cannot_change_assignee_403(self):
        """owner 不能通过行内编辑修改负责人（仅管理员）。"""
        self._login('owner1', 'owner123456')
        resp = self._post_field(self.task_id, 'assignee', self.owner2_id)
        self.assertEqual(resp.status_code, 403)
        self.assertFalse(resp.get_json()['success'])
        self.assertEqual(models.get_task(self.task_id)['assignee'], self.owner1_id)


# ============================================================
# 11. V2 过程证据 / 阻塞记录：CRUD 与权限
# ============================================================

class TestEvidenceBlockers(unittest.TestCase):
    """V2 证据/阻塞 CRUD：添加（owner 自己任务）、删除（仅 admin）、越权/错位/不存在 403/404。"""

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()
        self.client = make_client(self.app)
        self.admin_id = create_test_user('admin', '管理员', password='admin123456', role='admin')
        self.owner1_id = create_test_user('owner1', '张三', password='owner123456', role='owner')
        self.owner2_id = create_test_user('owner2', '李四', password='owner123456', role='owner')
        # admin 创建任务指派给 owner1
        self.task1 = create_task_direct('owner1的任务', self.admin_id, self.owner1_id)
        self.task2 = create_task_direct('owner2的任务', self.admin_id, self.owner2_id)

    def _login(self, username, password='owner123456'):
        return self.client.post('/login', data={
            'username': username, 'password': password,
        }, follow_redirects=False)

    def _add_evidence(self, task_id, etype='text', content='证据内容', **extra):
        return self.client.post('/tasks/%d/evidence' % task_id,
                                data=dict(etype=etype, content=content, **extra))

    def test_add_evidence_own_task(self):
        """owner 给自己的任务添加证据：成功且落库。"""
        self._login('owner1')
        resp = self._add_evidence(self.task1, etype='link', content='http://example.com/report')
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.get_json()['success'])

        evidence = models.get_evidence_list(self.task1)
        self.assertEqual(len(evidence), 1)
        self.assertEqual(evidence[0]['etype'], 'link')
        self.assertEqual(evidence[0]['content'], 'http://example.com/report')
        self.assertEqual(evidence[0]['created_by'], self.owner1_id)

    def test_add_evidence_others_task_403(self):
        """owner 给他人任务添加证据应返回 403。"""
        self._login('owner1')
        resp = self._add_evidence(self.task2, etype='text', content='越权证据')
        self.assertEqual(resp.status_code, 403)
        self.assertEqual(len(models.get_evidence_list(self.task2)), 0)

    def test_add_evidence_invalid_input(self):
        """添加证据：非法类型/空内容 400，任务不存在 404。"""
        self._login('owner1')

        resp = self._add_evidence(self.task1, etype='video', content='视频')
        self.assertEqual(resp.status_code, 400)
        self.assertFalse(resp.get_json()['success'])

        resp = self._add_evidence(self.task1, etype='text', content='   ')
        self.assertEqual(resp.status_code, 400)

        resp = self._add_evidence(99999, etype='text', content='幽灵')
        self.assertEqual(resp.status_code, 404)

    def test_add_blocker_and_resolve_flow(self):
        """阻塞记录：添加为 open；创建者/admin 可标记解决；重复解决 400。"""
        self._login('owner1')
        resp = self.client.post('/tasks/%d/blockers' % self.task1,
                                data={'content': '等待第三方接口开放'})
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.get_json()['success'])

        blockers = models.get_blockers(self.task1)
        self.assertEqual(len(blockers), 1)
        self.assertEqual(blockers[0]['status'], 'open')
        blocker_id = blockers[0]['blocker_id']

        # 创建者本人标记解决
        resp = self.client.post(
            '/tasks/%d/blockers/%d/resolve' % (self.task1, blocker_id))
        self.assertEqual(resp.status_code, 200)
        resolved = models.get_blockers(self.task1)[0]
        self.assertEqual(resolved['status'], 'resolved')
        self.assertEqual(resolved['resolved_by'], self.owner1_id)

        # 重复标记解决 → 400
        resp = self.client.post(
            '/tasks/%d/blockers/%d/resolve' % (self.task1, blocker_id))
        self.assertEqual(resp.status_code, 400)

        # admin 可解决他人创建的阻塞
        resp = self.client.post('/tasks/%d/blockers' % self.task1,
                                data={'content': '第二个阻塞'})
        blocker2 = [b for b in models.get_blockers(self.task1) if b['status'] == 'open'][0]
        self.client.get('/logout')
        self._login('admin', 'admin123456')
        resp = self.client.post(
            '/tasks/%d/blockers/%d/resolve' % (self.task1, blocker2['blocker_id']))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            models.get_blockers(self.task1)[0]['status'], 'resolved')

    def test_blocker_resolve_forbidden_and_not_found(self):
        """标记解决：非创建者非 admin 403；阻塞不存在或不属于该任务 404。"""
        self._login('owner1')
        self.client.post('/tasks/%d/blockers' % self.task1, data={'content': '待解决阻塞'})
        blocker_id = models.get_blockers(self.task1)[0]['blocker_id']

        # owner2（非创建者、非 admin）尝试解决 → 403
        self.client.get('/logout')
        self._login('owner2')
        resp = self.client.post(
            '/tasks/%d/blockers/%d/resolve' % (self.task1, blocker_id))
        self.assertEqual(resp.status_code, 403)
        self.assertEqual(models.get_blockers(self.task1)[0]['status'], 'open',
                         'owner2 不应能解决他人创建的阻塞')

        # 不存在的阻塞 → 404
        resp = self.client.post('/tasks/%d/blockers/99999/resolve' % self.task1)
        self.assertEqual(resp.status_code, 404)

        # 阻塞属于 task1 但通过 task2 的 URL 访问 → 404
        self.client.get('/logout')
        self._login('admin', 'admin123456')
        resp = self.client.post(
            '/tasks/%d/blockers/%d/resolve' % (self.task2, blocker_id))
        self.assertEqual(resp.status_code, 404)

    def test_delete_evidence_permissions(self):
        """删除证据：admin 成功；owner 403；不属于该任务/不存在 404。"""
        self._login('owner1')
        self._add_evidence(self.task1, etype='text', content='待删除证据')
        evidence_id = models.get_evidence_list(self.task1)[0]['evidence_id']

        # owner 删除自己任务上的证据 → 403（删除仅 admin）
        resp = self.client.post(
            '/tasks/%d/evidence/%d/delete' % (self.task1, evidence_id),
            headers={'X-Requested-With': 'XMLHttpRequest'})
        self.assertEqual(resp.status_code, 403)
        self.assertEqual(len(models.get_evidence_list(self.task1)), 1,
                         'owner 删除被拒后证据应仍在')

        # admin 通过错误的 task URL 删除（证据属于 task1）→ 404
        self.client.get('/logout')
        self._login('admin', 'admin123456')
        resp = self.client.post(
            '/tasks/%d/evidence/%d/delete' % (self.task2, evidence_id),
            headers={'X-Requested-With': 'XMLHttpRequest'})
        self.assertEqual(resp.status_code, 404)
        self.assertFalse(resp.get_json()['success'])

        # 不存在的证据 ID → 404
        resp = self.client.post(
            '/tasks/%d/evidence/99999/delete' % self.task1,
            headers={'X-Requested-With': 'XMLHttpRequest'})
        self.assertEqual(resp.status_code, 404)

        # admin 正确删除 → 200，证据消失
        resp = self.client.post(
            '/tasks/%d/evidence/%d/delete' % (self.task1, evidence_id),
            headers={'X-Requested-With': 'XMLHttpRequest'})
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.get_json()['success'])
        self.assertEqual(len(models.get_evidence_list(self.task1)), 0)

    def test_delete_blocker_permissions(self):
        """删除阻塞：admin 成功；owner 403。"""
        self._login('owner1')
        self.client.post('/tasks/%d/blockers' % self.task1, data={'content': '待删除阻塞'})
        blocker_id = models.get_blockers(self.task1)[0]['blocker_id']

        # owner 删除 → 403
        resp = self.client.post(
            '/tasks/%d/blockers/%d/delete' % (self.task1, blocker_id),
            headers={'X-Requested-With': 'XMLHttpRequest'})
        self.assertEqual(resp.status_code, 403)
        self.assertEqual(len(models.get_blockers(self.task1)), 1)

        # admin 删除 → 200
        self.client.get('/logout')
        self._login('admin', 'admin123456')
        resp = self.client.post(
            '/tasks/%d/blockers/%d/delete' % (self.task1, blocker_id),
            headers={'X-Requested-With': 'XMLHttpRequest'})
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.get_json()['success'])
        self.assertEqual(len(models.get_blockers(self.task1)), 0)


# ============================================================
# 12. V2 时间线留痕：证据/阻塞操作写入 progress_logs
# ============================================================

class TestTimelineLogging(unittest.TestCase):
    """V2 时间线留痕：添加/删除证据、添加阻塞均应在 progress_logs 留下记录。"""

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()
        self.client = make_client(self.app)
        self.admin_id = create_test_user('admin', '管理员', password='admin123456', role='admin')
        self.owner_id = create_test_user('owner', '张三', password='owner123456', role='owner')
        self.task_id = create_task_direct('留痕测试任务', self.admin_id, self.owner_id)

    def _login(self, username, password):
        return self.client.post('/login', data={
            'username': username, 'password': password,
        }, follow_redirects=False)

    def test_evidence_add_and_delete_write_log(self):
        """添加证据与删除证据均写入进度日志（纯备注型，无状态变更）。"""
        self._login('owner', 'owner123456')
        self.client.post('/tasks/%d/evidence' % self.task_id,
                         data={'etype': 'text', 'content': '阶段汇报材料'})

        logs = models.get_progress_logs(self.task_id)
        add_logs = [l for l in logs if (l['progress_note'] or '').startswith('添加证据')]
        self.assertEqual(len(add_logs), 1, '添加证据应写入 1 条时间线记录')
        self.assertIn('阶段汇报材料', add_logs[0]['progress_note'])
        self.assertIsNone(add_logs[0]['status_from'], '留痕日志不应有状态变更')
        self.assertIsNone(add_logs[0]['status_to'])
        self.assertEqual(add_logs[0]['operator'], self.owner_id)

        # admin 删除证据 → 也留痕
        self.client.get('/logout')
        self._login('admin', 'admin123456')
        evidence_id = models.get_evidence_list(self.task_id)[0]['evidence_id']
        self.client.post('/tasks/%d/evidence/%d/delete' % (self.task_id, evidence_id),
                         headers={'X-Requested-With': 'XMLHttpRequest'})
        logs = models.get_progress_logs(self.task_id)
        del_logs = [l for l in logs if (l['progress_note'] or '').startswith('删除证据')]
        self.assertEqual(len(del_logs), 1, '删除证据应写入 1 条时间线记录')

    def test_blocker_add_writes_log(self):
        """添加阻塞记录写入进度日志。"""
        self._login('owner', 'owner123456')
        self.client.post('/tasks/%d/blockers' % self.task_id,
                         data={'content': '等待预算审批'})

        logs = models.get_progress_logs(self.task_id)
        blocker_logs = [l for l in logs if (l['progress_note'] or '').startswith('添加阻塞')]
        self.assertEqual(len(blocker_logs), 1, '添加阻塞应写入 1 条时间线记录')
        self.assertIn('等待预算审批', blocker_logs[0]['progress_note'])


# ============================================================
# 13. V2 抽屉路由：任务抽屉 / Owner 抽屉 / 消息抽屉
# ============================================================

class TestDrawerRoutes(unittest.TestCase):
    """V2 抽屉路由：HTML 片段返回、404/登录拦截。"""

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()
        self.client = make_client(self.app)
        self.admin_id = create_test_user('admin', '管理员', password='admin123456', role='admin')
        self.owner_id = create_test_user('owner', '张三', password='owner123456', role='owner')
        self.task_id = create_task_direct('抽屉测试任务', self.admin_id, self.owner_id)

    def _login(self, username, password):
        return self.client.post('/login', data={
            'username': username, 'password': password,
        }, follow_redirects=False)

    def test_task_drawer_fragment(self):
        """GET /tasks/<id>/drawer 返回纯 HTML 片段（无整页包裹），含任务标题。"""
        self._login('admin', 'admin123456')
        resp = self.client.get('/tasks/%d/drawer' % self.task_id)
        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode('utf-8')
        # 片段不应包含整页结构
        self.assertNotIn('<!DOCTYPE html>', html, '抽屉应返回片段而非整页')
        self.assertNotIn('topbar', html, '抽屉片段不应包含顶栏（base.html 包裹）')
        self.assertIn('抽屉测试任务', html)
        # 含 3 页签结构（详情/过程证据/阻塞记录）
        self.assertIn('过程证据', html)
        self.assertIn('阻塞', html)

    def test_owner_drawer_fragment(self):
        """GET /tasks/owner/<uid>/drawer 返回负责人任务列表片段（admin 可见推送提醒）。"""
        self._login('admin', 'admin123456')
        resp = self.client.get('/tasks/owner/%d/drawer' % self.owner_id)
        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode('utf-8')
        self.assertNotIn('<!DOCTYPE html>', html)
        self.assertIn('张三', html)
        self.assertIn('抽屉测试任务', html)
        # admin 视角应看到推送提醒按钮
        self.assertIn('推送提醒', html)
        self.assertIn('/tasks/remind', html)

    def test_message_drawer_fragment(self):
        """GET /messages/drawer 返回消息抽屉片段，含最近消息内容。"""
        # admin 创建任务指派给 owner → owner 收到指派消息
        self._login('admin', 'admin123456')
        self.client.post('/tasks/new', data={
            'title': '指派消息任务', 'assignee': self.owner_id,
            'priority': 'medium', 'due_date': make_future_date(7),
        }, follow_redirects=False)

        self.client.get('/logout')
        self._login('owner', 'owner123456')
        resp = self.client.get('/messages/drawer')
        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode('utf-8')
        self.assertNotIn('<!DOCTYPE html>', html)
        self.assertIn('指派消息任务', html)

    def test_drawer_negative_cases(self):
        """抽屉路由：未登录 302；任务/负责人不存在 404。"""
        # 未登录 → 302 到 login
        for url in ('/tasks/%d/drawer' % self.task_id,
                    '/tasks/owner/%d/drawer' % self.owner_id,
                    '/messages/drawer'):
            resp = self.client.get(url)
            self.assertEqual(resp.status_code, 302, '%s 未登录应 302' % url)
            self.assertIn('/login', resp.headers.get('Location', ''))

        # 登录后：不存在的任务/负责人 → 404
        self._login('admin', 'admin123456')
        resp = self.client.get('/tasks/99999/drawer')
        self.assertEqual(resp.status_code, 404)
        resp = self.client.get('/tasks/owner/99999/drawer')
        self.assertEqual(resp.status_code, 404)


# ============================================================
# 14. V2 推送提醒：POST /tasks/remind
# ============================================================

class TestRemind(unittest.TestCase):
    """V2 推送提醒：task_id / owner_id 两种方式消息落库；权限与参数校验。"""

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()
        self.client = make_client(self.app)
        self.admin_id = create_test_user('admin', '管理员', password='admin123456', role='admin')
        self.owner_id = create_test_user('owner', '张三', password='owner123456', role='owner')
        self.task_id = create_task_direct('提醒测试任务', self.admin_id, self.owner_id)

    def _login(self, username, password):
        return self.client.post('/login', data={
            'username': username, 'password': password,
        }, follow_redirects=False)

    def test_remind_by_task_id(self):
        """admin 按 task_id 推送提醒：任务负责人收到 admin_directive 消息。"""
        self._login('admin', 'admin123456')
        resp = self.client.post('/tasks/remind', data={'task_id': str(self.task_id)})
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.get_json()['success'])

        msgs = models.get_messages(self.owner_id)
        remind_msgs = [m for m in msgs if m['type'] == 'admin_directive']
        self.assertEqual(len(remind_msgs), 1, '负责人应收到 1 条督办提醒')
        self.assertIn('提醒测试任务', remind_msgs[0]['content'])
        self.assertEqual(remind_msgs[0]['task_id'], self.task_id)
        self.assertEqual(remind_msgs[0]['sender'], self.admin_id)

    def test_remind_by_owner_id(self):
        """admin 按 owner_id 推送提醒：该负责人收到泛化提醒消息。"""
        self._login('admin', 'admin123456')
        resp = self.client.post('/tasks/remind', data={'owner_id': str(self.owner_id)})
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.get_json()['success'])

        msgs = models.get_messages(self.owner_id)
        remind_msgs = [m for m in msgs if m['type'] == 'admin_directive']
        self.assertEqual(len(remind_msgs), 1)
        self.assertIn('督办提醒', remind_msgs[0]['content'])

    def test_remind_negative_cases(self):
        """推送提醒：owner 403；缺参数 400；任务不存在 404。"""
        # owner 推送 → 403
        self._login('owner', 'owner123456')
        resp = self.client.post('/tasks/remind', data={'task_id': str(self.task_id)})
        self.assertEqual(resp.status_code, 403)
        self.assertFalse(resp.get_json()['success'])
        self.assertEqual(len(models.get_messages(self.owner_id)), 0, '403 时不应产生消息')

        # admin 缺少参数 → 400
        self.client.get('/logout')
        self._login('admin', 'admin123456')
        resp = self.client.post('/tasks/remind', data={})
        self.assertEqual(resp.status_code, 400)

        # 任务不存在 → 404
        resp = self.client.post('/tasks/remind', data={'task_id': '99999'})
        self.assertEqual(resp.status_code, 404)


# ============================================================
# 15. V2 登录页与暗色模式
# ============================================================

class TestLoginV2(unittest.TestCase):
    """V2 登录页交互：记住我 cookie、字段级错误提示；暗色模式切换机制。"""

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()
        self.client = make_client(self.app)
        create_test_user('admin', '管理员', password='admin123456', role='admin')

    def _login(self, username='admin', password='admin123456', **extra):
        return self.client.post('/login', data=dict(
            username=username, password=password, **extra), follow_redirects=False)

    def test_dark_mode_toggle_present(self):
        """登录后页面应包含暗色模式切换机制（html.dark 类 + db-theme 持久化）。"""
        self._login()
        resp = self.client.get('/dashboard')
        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode('utf-8')
        self.assertIn('theme-toggle', html, '应有主题切换按钮')
        self.assertIn("classList.add('dark')", html, '应向 html 根元素挂 dark 类')
        self.assertIn('db-theme', html, '应用 localStorage 持久化主题偏好')

    def test_login_remember_me_cookie(self):
        """记住我：勾选写 30 天 cookie，登录页回填用户名；不勾选清除 cookie。"""
        # 勾选记住我 → Set-Cookie remember_username
        resp = self._login(remember='1')
        self.assertEqual(resp.status_code, 302)
        cookies = '; '.join(resp.headers.getlist('Set-Cookie'))
        self.assertIn('remember_username=admin', cookies, '勾选记住我应写入用户名 cookie')

        # 携带 cookie 访问登录页 → 用户名回填 + 复选框预勾选
        self.client.set_cookie('remember_username', 'admin')
        resp = self.client.get('/login')
        html = resp.data.decode('utf-8')
        self.assertIn('value="admin"', html, '登录页应回填记住的用户名')
        self.assertIn('checked', html, '记住我复选框应预勾选')

        # 不勾选记住我 → cookie 被清除
        self.client.delete_cookie('remember_username')
        resp = self._login()
        self.assertEqual(resp.status_code, 302)
        cookies = '; '.join(resp.headers.getlist('Set-Cookie'))
        self.assertNotIn('remember_username=admin', cookies, '不勾选时不应写用户名 cookie')
        self.assertIn('remember_username=;', cookies, '应清除已记住的用户名 cookie')

    def test_login_field_level_errors(self):
        """字段级错误提示：空用户名/空密码返回 400 并定位到对应字段。"""
        # 空用户名
        resp = self.client.post('/login', data={'username': '', 'password': 'xxx'},
                                follow_redirects=False)
        self.assertEqual(resp.status_code, 400)
        html = resp.data.decode('utf-8')
        self.assertIn('请输入用户名', html)
        self.assertIn('has-error', html)

        # 空密码
        resp = self.client.post('/login', data={'username': 'admin', 'password': ''},
                                follow_redirects=False)
        self.assertEqual(resp.status_code, 400)
        html = resp.data.decode('utf-8')
        self.assertIn('请输入密码', html)
        self.assertIn('has-error', html)

    def test_logout_flash_consumed_on_login_page(self):
        """登出提示必须在登录页消费，不能泄漏到登录后的首页。"""
        self._login()
        # 登出会 flash '您已安全退出' 并重定向到登录页
        resp = self.client.get('/logout', follow_redirects=False)
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp.headers.get('Location', ''))

        # 登录页应消费并显示该提示
        resp = self.client.get('/login', follow_redirects=False)
        self.assertEqual(resp.status_code, 200)
        login_html = resp.data.decode('utf-8')
        self.assertIn('您已安全退出', login_html)
        # 登录页含 flash 容器，表示已消费
        self.assertIn('flash-container', login_html)

        # 重新登录后，首页不应再出现该提示
        resp = self._login()
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/dashboard', resp.headers.get('Location', ''))
        resp = self.client.get('/dashboard', follow_redirects=False)
        self.assertEqual(resp.status_code, 200)
        dashboard_html = resp.data.decode('utf-8')
        self.assertNotIn('您已安全退出', dashboard_html,
                         '登出提示不应被带到登录后的 dashboard')


# ============================================================
# 16. V2 新建/编辑任务表单 3 字段：进度 / 风险点 / 协同方
# ============================================================

class TestTaskFormExtraFields(unittest.TestCase):
    """V2 表单 3 字段：progress_percent / risk_note / collaborators 保存与校验。"""

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()
        self.client = make_client(self.app)
        self.admin_id = create_test_user('admin', '管理员', password='admin123456', role='admin')
        self.owner_id = create_test_user('owner', '张三', role='owner')
        self.client.post('/login', data={
            'username': 'admin', 'password': 'admin123456',
        }, follow_redirects=False)

    def test_create_task_with_extra_fields_saved(self):
        """新建任务携带 3 字段：保存成功且值正确落库。"""
        resp = self.client.post('/tasks/new', data={
            'title': '三字段任务',
            'description': '',
            'assignee': self.owner_id,
            'priority': 'high',
            'due_date': make_future_date(7),
            'progress_percent': '30',
            'risk_note': '存在外部供应商交付风险',
            'collaborators': '财务部、采购部',
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)

        task = models.get_task(1)
        self.assertIsNotNone(task)
        self.assertEqual(task['progress_percent'], 30)
        self.assertEqual(task['risk_note'], '存在外部供应商交付风险')
        self.assertEqual(task['collaborators'], '财务部、采购部')

    def test_create_task_progress_150_rejected(self):
        """新建任务进度填 150：拒绝创建并提示错误。"""
        resp = self.client.post('/tasks/new', data={
            'title': '非法进度任务',
            'description': '',
            'assignee': self.owner_id,
            'priority': 'medium',
            'due_date': make_future_date(7),
            'progress_percent': '150',
        }, follow_redirects=False)
        # 校验失败重新渲染表单（200），不创建任务
        self.assertEqual(resp.status_code, 200)
        self.assertIn('进度必须在 0-100 之间', resp.data.decode('utf-8'))
        self.assertIsNone(models.get_task(1), '进度越界的任务不应被创建')

    def test_edit_task_extra_fields_saved(self):
        """编辑任务可更新 3 字段。"""
        self.client.post('/tasks/new', data={
            'title': '原任务', 'assignee': self.owner_id,
            'priority': 'medium', 'due_date': make_future_date(7),
        }, follow_redirects=False)

        resp = self.client.post('/tasks/1/edit', data={
            'title': '更新后任务',
            'description': '',
            'assignee': self.owner_id,
            'priority': 'medium',
            'due_date': make_future_date(14),
            'progress_percent': '80',
            'risk_note': '风险已升级',
            'collaborators': '技术部',
        }, follow_redirects=False)
        self.assertEqual(resp.status_code, 302)

        task = models.get_task(1)
        self.assertEqual(task['progress_percent'], 80)
        self.assertEqual(task['risk_note'], '风险已升级')
        self.assertEqual(task['collaborators'], '技术部')

    def test_new_task_date_input_shows_text_placeholder(self):
        """新建页截止日期默认以 text 类型显示自定义 placeholder，聚焦后切换为 date。"""
        resp = self.client.get('/tasks/new')
        self.assertEqual(resp.status_code, 200)
        html = resp.data.decode('utf-8')
        # 未填值时 type="text"，以便显示自定义 placeholder
        self.assertIn('type="text"', html)
        self.assertIn('placeholder="点击选择截止日期"', html)
        # 聚焦/失焦切换 type 的内联事件
        self.assertIn("onfocus=\"if(this.type!=='date'){this.type='date';}\"", html)
        self.assertIn("onblur=\"if(!this.value){this.type='text';}\"", html)
        # 不应再使用旧的 wrapper hack
        self.assertNotIn('data-placeholder="点击选择截止日期"', html)
        self.assertNotIn('data-display-value', html)
        self.assertNotIn('<span class="date-placeholder">', html)


# ============================================================
# DEF-005：登录限流与密码强度校验
# ============================================================

class TestLoginSecurity(unittest.TestCase):
    """DEF-005：登录失败限流 + 弱口令拦截。

    阈值在 setUp 里调小（3 次 / 3 秒），让完整的「失败 → 锁定 → 拒绝
    → 到期解锁」周期能在秒级内跑完；tearDown 必须还原配置并清空计数，
    否则会污染后续测试类。
    """

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()
        auth.reset_login_attempts()
        self._saved = (config.LOGIN_MAX_ATTEMPTS,
                       config.LOGIN_LOCKOUT_MINUTES,
                       config.PASSWORD_MIN_LENGTH)
        config.LOGIN_MAX_ATTEMPTS = 3
        config.LOGIN_LOCKOUT_MINUTES = 0.05      # 3 秒
        config.PASSWORD_MIN_LENGTH = 8

        self.client = make_client(self.app)
        create_test_user('admin', '管理员', password='Admin@12345', role='admin')

    def tearDown(self):
        (config.LOGIN_MAX_ATTEMPTS,
         config.LOGIN_LOCKOUT_MINUTES,
         config.PASSWORD_MIN_LENGTH) = self._saved
        auth.reset_login_attempts()

    def _login(self, password='Admin@12345', username='admin', ip='127.0.0.1'):
        return self.client.post(
            '/login', data={'username': username, 'password': password},
            environ_base={'REMOTE_ADDR': ip}, follow_redirects=False)

    # ---------- 密码强度规则 ----------

    def test_weak_password_rejected_by_length(self):
        """短于 PASSWORD_MIN_LENGTH 的密码不合规。"""
        ok, errs = auth.check_password_strength('Ab1', 'someone')
        self.assertFalse(ok)
        self.assertTrue(any('至少 8 个字符' in e for e in errs))

    def test_weak_password_rejected_all_digits(self):
        """纯数字密码不合规（挡 12345678 这类长度够但强度为零的）。"""
        ok, errs = auth.check_password_strength('12345678')
        self.assertFalse(ok)
        self.assertIn('密码不能全是数字', errs)

    def test_weak_password_rejected_all_letters(self):
        """纯字母密码不合规。"""
        ok, errs = auth.check_password_strength('abcdefghij')
        self.assertFalse(ok)
        self.assertIn('密码不能全是字母，请混合数字', errs)

    def test_weak_password_rejected_by_blacklist(self):
        """常见弱口令即使长度达标也拦截，且比较不区分大小写。"""
        for pwd in ('password123', 'Admin1234', 'QWERTY123'):
            ok, errs = auth.check_password_strength(pwd)
            self.assertFalse(ok, f'{pwd} 应被弱口令黑名单拦截')
            self.assertIn('该密码过于常见，容易被猜到，请换一个', errs)

    def test_weak_password_rejected_same_as_username(self):
        """密码不能与用户名相同。"""
        ok, errs = auth.check_password_strength('zhangsan88', 'zhangsan88')
        self.assertFalse(ok)
        self.assertIn('密码不能与用户名相同', errs)

    def test_strong_password_accepted(self):
        """字母+数字混合、不在黑名单内 → 合规。"""
        ok, errs = auth.check_password_strength('Admin@12345')
        self.assertTrue(ok, f'不应有错误: {errs}')
        self.assertEqual(errs, [])

    def test_strength_check_can_be_disabled(self):
        """PASSWORD_REQUIRE_STRENGTH=False 时只保留长度校验。"""
        config.PASSWORD_REQUIRE_STRENGTH = False
        try:
            ok, _ = auth.check_password_strength('12345678')   # 纯数字但长度达标
            self.assertTrue(ok)
        finally:
            config.PASSWORD_REQUIRE_STRENGTH = True

    # ---------- 登录失败限流 ----------

    def test_failure_counter_decrements(self):
        """每次失败提示的剩余次数递减，让真实用户知道还剩几次机会。"""
        r1 = self._login(password='wrong1')
        self.assertEqual(r1.status_code, 401)
        self.assertIn('还可尝试 2 次', r1.data.decode('utf-8'))

        r2 = self._login(password='wrong2')
        self.assertEqual(r2.status_code, 401)
        self.assertIn('还可尝试 1 次', r2.data.decode('utf-8'))

    def test_lockout_triggered_after_max_attempts(self):
        """达到阈值后账号被临时锁定，且当次就告知要等多久。"""
        resp = None
        for i in range(3):
            resp = self._login(password=f'wrong{i}')
        body = resp.data.decode('utf-8')
        self.assertIn('临时锁定', body)
        self.assertIn('分钟后再试', body, '触发锁定当次应给出等待时长')

    def test_lockout_rejects_even_correct_password(self):
        """锁定期内即使密码正确也拒绝——否则限流形同虚设。"""
        for i in range(3):
            self._login(password=f'wrong{i}')

        resp = self._login(password='Admin@12345')
        self.assertEqual(resp.status_code, 401, '锁定期内正确密码也必须被拒')
        body = resp.data.decode('utf-8')
        self.assertIn('登录失败次数过多', body)
        self.assertIn('后再试', body)

    def test_lockout_expires_automatically(self):
        """锁定期过后自动解锁，无需人工干预。"""
        for i in range(3):
            self._login(password=f'wrong{i}')
        time.sleep(3.5)          # LOGIN_LOCKOUT_MINUTES=0.05 → 3 秒

        self.assertEqual(self._login(password='Admin@12345').status_code, 302,
                         '到期后应恢复登录')

    def test_successful_login_clears_counter(self):
        """登录成功后计数清零，不会带着旧账继续累加。"""
        self._login(password='wrong1')
        self._login(password='wrong2')            # 再错一次就锁
        self.assertEqual(self._login(password='Admin@12345').status_code, 302)

        body = self._login(password='wrong1').data.decode('utf-8')
        self.assertIn('还可尝试 2 次', body, '成功后应重新从满额开始计数')

    def test_lockout_isolated_by_ip(self):
        """同一用户名在不同 IP 下分别计数。"""
        for i in range(3):
            self._login(password=f'wrong{i}', ip='10.0.0.9')
        self.assertEqual(
            self._login(password='Admin@12345', ip='10.0.0.10').status_code, 302,
            '另一 IP 不应被牵连')

    def test_lockout_isolated_by_username(self):
        """同一 IP 下不同用户名分别计数（避免共用出口 IP 时互相误伤）。"""
        create_test_user('owner9', '负责人九', password='Owner@12345', role='owner')
        for i in range(3):
            self._login(password=f'wrong{i}', username='admin')
        self.assertEqual(
            self._login(password='Owner@12345', username='owner9').status_code, 302,
            'admin 被锁不应影响同 IP 的其他账号')

    # ---------- 弱口令在 4 个业务入口被拦截 ----------

    def _login_as_admin(self):
        self.assertEqual(self._login(password='Admin@12345').status_code, 302)

    def test_create_user_rejects_weak_password(self):
        """新建用户：弱口令被拒且不落库。"""
        self._login_as_admin()
        self.client.post('/users/new', data={
            'username': 'weakman', 'display_name': '弱口令',
            'password': '123456', 'role': 'owner'})
        self.assertIsNone(models.get_user_by_username('weakman'))

    def test_create_user_accepts_strong_password(self):
        """新建用户：强口令正常创建。"""
        self._login_as_admin()
        self.client.post('/users/new', data={
            'username': 'strongman', 'display_name': '强口令',
            'password': 'Good@2026x', 'role': 'owner'})
        self.assertIsNotNone(models.get_user_by_username('strongman'))

    def test_reset_password_rejects_weak_password(self):
        """重置密码：弱口令不得改写密码哈希。"""
        self._login_as_admin()
        uid = create_test_user('owner8', '负责人八', password='Owner@12345', role='owner')
        before = models.get_user(uid)['password_hash']
        self.client.post(f'/users/{uid}/reset-password',
                         data={'new_password': '12345678'})
        self.assertEqual(models.get_user(uid)['password_hash'], before)

    def test_profile_change_rejects_weak_password(self):
        """个人改密码：弱口令不得改写密码哈希。"""
        self._login_as_admin()
        before = models.get_user_by_username('admin')['password_hash']
        self.client.post('/settings/profile', data={
            'display_name': '管理员', 'old_password': 'Admin@12345',
            'new_password': 'abcdefgh', 'confirm_password': 'abcdefgh'})
        self.assertEqual(models.get_user_by_username('admin')['password_hash'],
                         before)

    def test_existing_weak_password_account_can_still_login(self):
        """不追溯：历史遗留的弱口令账号仍能登录，只在下次改密码时才要求达标。"""
        create_test_user('oldman', '老账号', password='123456', role='owner')
        self.assertEqual(self._login(password='123456', username='oldman').status_code, 302)


class TestBehindProxy(unittest.TestCase):
    """反向代理下的客户端 IP 透传。

    登录限流按「IP + 用户名」计数，IP 取错会让限流彻底失效或误伤一片：
      - 不开透传 → 所有请求都记成代理 IP，一人被锁、全员陪绑（见反证用例）；
      - 默认开启透传 → 任何人都能伪造 X-Forwarded-For 绕过限流。
    因此 BEHIND_PROXY 默认关闭，只在确实有反向代理时才打开。
    """

    def _make_app(self, behind_proxy, hops=1):
        """建一个带 /_probe_ip 探针的 app，用来观察应用实际看到的客户端 IP。

        必须用 test_client 发真实请求：ProxyFix 是 WSGI 中间件，
        test_request_context 会绕过 wsgi_app，测不出效果。
        """
        config.BEHIND_PROXY = behind_proxy
        config.BEHIND_PROXY_TRUSTED_HOPS = hops
        a = create_app()
        a.config['TESTING'] = True

        @a.route('/_probe_ip')
        def _probe_ip():
            return request.remote_addr

        return a

    def setUp(self):
        reset_database()
        auth.reset_login_attempts()
        self._saved = (config.BEHIND_PROXY, config.BEHIND_PROXY_TRUSTED_HOPS,
                       config.LOGIN_MAX_ATTEMPTS, config.LOGIN_LOCKOUT_MINUTES)
        config.LOGIN_MAX_ATTEMPTS = 3
        config.LOGIN_LOCKOUT_MINUTES = 15
        self.app = self._make_app(False)
        self.client = make_client(self.app)
        create_test_user('admin', '管理员', password='Admin@12345', role='admin')

    def tearDown(self):
        (config.BEHIND_PROXY, config.BEHIND_PROXY_TRUSTED_HOPS,
         config.LOGIN_MAX_ATTEMPTS, config.LOGIN_LOCKOUT_MINUTES) = self._saved
        auth.reset_login_attempts()

    def _login(self, ip, password='Admin@12345', client=None):
        return (client or self.client).post(
            '/login', data={'username': 'admin', 'password': password},
            headers={'X-Forwarded-For': ip}, follow_redirects=False)

    def test_spoofed_header_ignored_by_default(self):
        """默认不信任 X-Forwarded-For，否则人人都能伪造 IP 绕过限流。"""
        ip = self.client.get(
            '/_probe_ip', headers={'X-Forwarded-For': '203.0.113.9'}).data.decode()
        self.assertEqual(ip, '127.0.0.1')

    def test_forwarded_header_used_when_enabled(self):
        """开启 BEHIND_PROXY 后取代理透传的真实 IP；无代理头时回落到直连 IP。"""
        c = make_client(self._make_app(True))
        self.assertEqual(
            c.get('/_probe_ip',
                  headers={'X-Forwarded-For': '203.0.113.9'}).data.decode(),
            '203.0.113.9')
        self.assertEqual(c.get('/_probe_ip').data.decode(), '127.0.0.1',
                         '没有代理头时应回落到直连 IP，不能崩')

    def test_only_configured_hops_trusted(self):
        """信任 1 层时取倒数第一跳，与 BEHIND_PROXY_TRUSTED_HOPS 语义一致。"""
        c = make_client(self._make_app(True, hops=1))
        ip = c.get('/_probe_ip',
                   headers={'X-Forwarded-For': '203.0.113.9, 198.51.100.7'}
                   ).data.decode()
        self.assertEqual(ip, '198.51.100.7')

    def test_lockout_not_shared_across_real_clients(self):
        """开启透传后，一个人的失败不会把其他客户端一起锁掉。"""
        c = make_client(self._make_app(True))
        for i in range(3):
            self._login('203.0.113.9', password=f'bad{i}', client=c)

        self.assertEqual(self._login('203.0.113.9', client=c).status_code, 401,
                         '该客户端应被锁（正确密码也拒）')
        self.assertEqual(self._login('198.51.100.7', client=c).status_code, 302,
                         '另一客户端不应被牵连')

    def test_lockout_shared_when_transparent_proxy_disabled(self):
        """反证：不开 BEHIND_PROXY 时，不同来源会被记成同一 IP 而互相牵连。

        这条用例存在的意义是记录「为什么要有这个开关」——
        它演示的正是反向代理场景下不开透传会踩到的坑。
        """
        for i in range(3):
            self._login('203.0.113.9', password=f'bad{i}')
        self.assertEqual(self._login('198.51.100.7').status_code, 401,
                         '未开透传时所有人共用代理 IP，会被一起锁住')


# ============================================================
# DEF-002：CSRF 防护
# ============================================================

class TestCSRF(unittest.TestCase):
    """全站 CSRF 防护测试。

    除了验证令牌本身，这里还放了两条**防回归**的用例：
      - test_all_write_routes_require_token：遍历 url_map，确保每一条写路由
        都被保护住。将来新加 POST 路由却忘了防护，这条会直接失败。
      - test_every_post_form_has_token_field：扫描模板，确保每个 method=POST
        的表单都带了 hidden 字段。漏一个表单，用户提交时就会撞上 400。
    """

    # CSRF 拦截页的固定文案，用来区分「被安全机制拦下」和「路由自己返回 400」
    MARKER = '请求未通过安全校验'

    def setUp(self):
        reset_database()
        auth.reset_login_attempts()
        self.app = create_app()
        self.app.config['TESTING'] = True

        # 探针路由：一个受保护、一个豁免，用来验证机制本身
        @self.app.route('/_csrf_probe', methods=['POST'])
        def _csrf_probe():
            return {'ok': True}, 200

        @self.app.route('/_csrf_open', methods=['POST'])
        @csrf.csrf_exempt
        def _csrf_open():
            return {'ok': True}, 200

        create_test_user('admin', '管理员', password='Admin@12345', role='admin')
        self.raw = self.app.test_client()      # 不带令牌的原始客户端
        self.client = make_client(self.app)    # 自动带令牌

    def tearDown(self):
        auth.reset_login_attempts()

    def _token_from_page(self, client=None, path='/login'):
        """从渲染出的页面里抠出表单里的 CSRF 令牌。"""
        html = (client or self.raw).get(path).get_data(as_text=True)
        m = re.search(r'name="csrf_token" value="([^"]+)"', html)
        return m.group(1) if m else None

    # ---------- 基本行为 ----------

    def test_write_without_token_rejected(self):
        """不带令牌的写请求一律拒绝。"""
        self.assertEqual(self.raw.post('/_csrf_probe').status_code, 400)

    def test_write_with_valid_token_passes(self):
        """带上有效令牌的写请求正常放行。"""
        self.assertEqual(self.client.post('/_csrf_probe').status_code, 200)

    def test_forged_token_rejected(self):
        """伪造的令牌应被拒绝。"""
        r = self.raw.post('/_csrf_probe', data={'csrf_token': '9999999999.' + 'a' * 64})
        self.assertEqual(r.status_code, 400)

    def test_empty_token_rejected(self):
        """空令牌等同于没有令牌。"""
        self.assertEqual(
            self.raw.post('/_csrf_probe', data={'csrf_token': ''}).status_code, 400)

    def test_get_request_not_checked(self):
        """只读请求不做校验，否则正常浏览都会被拦。"""
        self.assertEqual(self.raw.get('/login').status_code, 200)

    def test_exempt_route_allowed(self):
        """被 csrf_exempt 豁免的路由不校验。"""
        self.assertEqual(self.raw.post('/_csrf_open').status_code, 200)

    def test_header_and_form_and_json_all_accepted(self):
        """三种携带方式都应被接受：表单字段 / 请求头 / JSON body。"""
        tok = self._token_from_page()
        self.assertEqual(
            self.raw.post('/_csrf_probe', data={'csrf_token': tok}).status_code, 200)
        self.assertEqual(
            self.raw.post('/_csrf_probe',
                          headers={'X-CSRF-Token': self._token_from_page()}).status_code, 200)
        self.assertEqual(
            self.raw.post('/_csrf_probe',
                          json={'csrf_token': self._token_from_page()}).status_code, 200)

    def test_token_not_transferable_to_another_session(self):
        """A 会话的令牌拿到 B 会话用应失效——否则攻击者自己开个页面就够了。"""
        tok = self._token_from_page()
        other = self.app.test_client()   # 全新会话
        r = other.post('/_csrf_probe', data={'csrf_token': tok})
        self.assertEqual(r.status_code, 400)

    def test_expired_token_rejected(self):
        """超过 CSRF_TOKEN_MAX_AGE 的令牌应失效。"""
        real = csrf.time.time
        try:
            csrf.time.time = lambda: real() - (config.CSRF_TOKEN_MAX_AGE + 60)
            with self.app.test_request_context('/login'):
                expired = csrf.generate_csrf_token()
                csrf.time.time = real
                self.assertFalse(csrf.validate_csrf_token(expired))
        finally:
            csrf.time.time = real

    def test_token_rotated_on_login(self):
        """登录成功后轮换会话密钥，登录前的令牌应作废。"""
        tok = self._token_from_page()
        r = self.raw.post('/login', data={'username': 'admin',
                                          'password': 'Admin@12345',
                                          'csrf_token': tok})
        self.assertEqual(r.status_code, 302, '登录应成功跳转')
        self.assertEqual(
            self.raw.post('/_csrf_probe', data={'csrf_token': tok}).status_code, 400,
            '登录轮换密钥后，登录前签发的令牌不该还能用')

    def test_disabled_switch_allows_untokened_write(self):
        """CSRF_ENABLED=False 时完全不校验（排障通道）。"""
        config.CSRF_ENABLED = False
        try:
            self.assertEqual(self.raw.post('/_csrf_probe').status_code, 200)
        finally:
            config.CSRF_ENABLED = True

    # ---------- 失败时的响应形态 ----------

    def test_ajax_failure_returns_json(self):
        """AJAX 请求失败应拿到 JSON，而不是一整页 HTML。"""
        r = self.raw.post('/_csrf_probe', headers={'X-Requested-With': 'XMLHttpRequest'})
        self.assertEqual(r.status_code, 400)
        self.assertTrue(r.is_json, '前端要解析 JSON，返回 HTML 会直接抛异常')
        self.assertIn('安全校验', r.get_json()['message'])

    def test_page_failure_returns_400_page(self):
        """普通表单提交失败应看到一个能读懂的 400 页面。"""
        r = self.raw.post('/_csrf_probe')
        text = r.get_data(as_text=True)
        self.assertEqual(r.status_code, 400)
        self.assertIn(self.MARKER, text)

    # ---------- 防回归 ----------

    def test_all_write_routes_require_token(self):
        """每一条写路由都必须被保护住（新增路由漏防护会在这里暴露）。"""
        unsafe = {'POST', 'PUT', 'PATCH', 'DELETE'}
        unchecked = []
        total = 0

        for rule in self.app.url_map.iter_rules():
            if not (set(rule.methods or ()) & unsafe):
                continue
            if rule.endpoint == 'static':
                continue
            # 显式豁免过的路由是「有意为之」，不算漏防护
            if csrf.is_exempt(self.app.view_functions.get(rule.endpoint)):
                continue
            url = re.sub(r'<[^>]+>', '1', str(rule))
            total += 1
            r = self.raw.post(url)
            if self.MARKER not in r.get_data(as_text=True):
                unchecked.append(f'{rule.endpoint} ({url}) -> {r.status_code}')

        self.assertGreater(total, 15, '写路由数量太少，url_map 可能没注册全')
        self.assertEqual(unchecked, [],
                         '以下写路由没有 CSRF 保护:\n  ' + '\n  '.join(unchecked))

    def test_every_post_form_has_token_field(self):
        """每个 method=POST 的表单都必须带 hidden 令牌字段。"""
        tpl_root = os.path.join(SRC_DIR, 'templates')
        form_re = re.compile(r'<form\b[^>]*>(.*?)</form>', re.IGNORECASE | re.DOTALL)
        missing = []
        total = 0

        for dirpath, _, filenames in os.walk(tpl_root):
            for fn in filenames:
                if not fn.endswith('.html'):
                    continue
                path = os.path.join(dirpath, fn)
                with open(path, 'r', encoding='utf-8') as f:
                    text = f.read()
                for m in form_re.finditer(text):
                    tag = text[m.start():text.find('>', m.start()) + 1]
                    if not re.search(r'method\s*=\s*["\']POST["\']', tag, re.IGNORECASE):
                        continue
                    total += 1
                    if 'name="csrf_token"' not in m.group(1):
                        missing.append(os.path.relpath(path, tpl_root))

        self.assertGreater(total, 15, '扫到的 POST 表单太少，模板路径可能不对')
        self.assertEqual(missing, [], '以下模板的 POST 表单缺少 CSRF 字段: '
                                      + ', '.join(sorted(set(missing))))

    def test_mainjs_injects_token_into_every_write_fetch(self):
        """main.js 里每个 POST 型 fetch 都必须经过 csrfHeaders。"""
        js_path = os.path.join(SRC_DIR, 'static', 'js', 'main.js')
        with open(js_path, 'r', encoding='utf-8') as f:
            js = f.read()

        self.assertIn('X-CSRF-Token', js, 'main.js 应统一注入 X-CSRF-Token 头')

        bad = []
        for m in re.finditer(r'fetch\s*\(', js):
            chunk = js[m.start():m.start() + 400]
            if "method: 'POST'" not in chunk:
                continue
            if 'csrfHeaders' not in chunk:
                bad.append(chunk.split('\n')[0][:60])
        self.assertEqual(bad, [], '以下 AJAX 写请求没有带 CSRF 令牌: ' + '; '.join(bad))


# ============================================================
# 20. V4 邮件设置页：配置、状态、记录、失败重发
# ============================================================

# 连本机 1 号端口：连接会被立刻拒绝，既不会真的发出邮件，
# 也不会像解析不存在的域名那样去等 DNS 超时。
MAIL_CFG_FORM = {
    'enabled': '1',
    'smtp_host': '127.0.0.1',
    'smtp_port': '1',
    'smtp_username': 'notice@example.com',
    'smtp_password': 'smtp-pass-123',
    'use_ssl': '1',
    'from_addr': 'notice@example.com',
    'from_name': '督办系统',
    'footer': '本邮件由督办系统自动发送',
    'batch_limit': '20',
    'retry_max': '3',
    'manual_cooldown': '300',
}


class TestMailSettings(unittest.TestCase):
    """V4 邮件设置页：配置保存与校验、状态展示、记录查询、失败重发。"""

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()
        self.client = make_client(self.app)
        self.admin_id = create_test_user('admin', '管理员',
                                         password='admin123456', role='admin')
        self.owner_id = create_test_user('owner', '张三',
                                         password='owner123456', role='owner')
        self._login('admin', 'admin123456')

    def _login(self, username, password):
        return self.client.post(
            '/login', data={'username': username, 'password': password},
            follow_redirects=False)

    def _html(self, resp):
        return resp.data.decode('utf-8')

    def _save(self, **overrides):
        """提交配置表单（可在 MAIL_CFG_FORM 基础上覆盖字段）。

        值为 None 表示「该字段不出现在表单里」——真实浏览器里未勾选的
        复选框根本不会提交，而路由侧判断的是键是否存在（不是值），
        所以这里必须真的把键删掉；传空字符串只会让键继续存在，
        模拟不出取消勾选的效果。
        """
        data = dict(MAIL_CFG_FORM)
        for key, value in overrides.items():
            if value is None:
                data.pop(key, None)
            else:
                data[key] = value
        return self.client.post('/mail/config', data=data, follow_redirects=True)

    def _log(self, recipient_id, email, mail_type, subject, success=1, error=None):
        """直接落一条发送历史。"""
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        return db.execute(
            "INSERT INTO email_log (recipient_id, recipient_email, task_id, mail_type,"
            " subject, operator_id, success, error_message, attempts, created_at,"
            " finished_at) VALUES (?,?,NULL,?,?,?,?,?,1,?,?)",
            (recipient_id, email, mail_type, subject, self.admin_id,
             success, error, now, now))

    # --- 权限 ---

    def test_status_page_requires_admin(self):
        """普通用户访问 /mail 应被拒（P-1 配置仅管理员可见）。"""
        self.client.get('/logout')
        self._login('owner', 'owner123456')
        self.assertEqual(self.client.get('/mail').status_code, 403)

    def test_status_page_requires_login(self):
        """未登录访问 /mail 应跳转登录。"""
        self.client.get('/logout')
        resp = self.client.get('/mail')
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp.headers.get('Location', ''))

    # --- 未配置时的展示 ---

    def test_unconfigured_shows_banner_with_reason(self):
        """未配置时页面顶部常驻红条，且给出具体原因（B5-②）。"""
        html = self._html(self.client.get('/mail'))
        self.assertIn('邮件功能当前不可用', html)
        self.assertIn('邮件功能未启用', html)   # 默认 MAIL_ENABLED=False

    def test_default_mail_disabled(self):
        """铁律：不配置 = 行为不变。默认必须关闭，否则离线包会去连 SMTP。"""
        self.assertFalse(models.get_mail_config()['enabled'])
        self.assertFalse(models.is_mail_configured())

    # --- 配置保存 ---

    def test_save_config_persists(self):
        """保存后各项应能读回。"""
        self._save()
        cfg = models.get_mail_config()
        self.assertEqual(cfg['smtp_host'], '127.0.0.1')
        self.assertEqual(cfg['smtp_port'], 1)
        self.assertEqual(cfg['from_addr'], 'notice@example.com')
        self.assertTrue(cfg['use_ssl'])
        self.assertTrue(cfg['enabled'])
        self.assertTrue(models.is_mail_configured())

    def test_password_stored_encrypted(self):
        """密码必须加密入库，且能解回原文（P-9）。"""
        self._save()
        stored = models.get_config('mail_smtp_password', '')
        self.assertNotIn('smtp-pass-123', stored, '密码不应以明文落库')
        self.assertEqual(models.get_mail_config()['smtp_password'], 'smtp-pass-123')

    def test_password_blank_keeps_previous(self):
        """密码框留空 = 不修改（否则每次保存配置都要重填授权码）。"""
        self._save()
        self._save(smtp_password='', smtp_host='smtp.other.com')
        self.assertEqual(models.get_mail_config()['smtp_password'], 'smtp-pass-123')

    def test_password_never_rendered(self):
        """页面上绝不能出现明文密码（P-7）。"""
        self._save()
        html = self._html(self.client.get('/mail'))
        self.assertNotIn('smtp-pass-123', html)

    def test_unchecking_checkbox_persists(self):
        """回归：复选框未勾选时不出现在 form 里，服务端必须显式补 0。

        否则「取消启用」会因为键缺失而保存不进去，开关卡死在开启状态。
        """
        self._save()
        self.assertTrue(models.get_mail_config()['enabled'])
        self._save(enabled=None)   # 模拟未勾选：键不出现在 form 里
        self.assertFalse(models.get_mail_config()['enabled'])

    def test_invalid_port_rejected(self):
        """端口越界应拒绝且不污染已存配置。"""
        self._save()
        resp = self._save(smtp_port='99999')
        self.assertIn('SMTP 端口应在 1-65535 之间', self._html(resp))
        self.assertEqual(models.get_mail_config()['smtp_host'], '127.0.0.1')

    def test_ssl_and_tls_mutex(self):
        """SSL 与 STARTTLS 同时勾选应被拒。"""
        resp = self._save(use_ssl='1', use_tls='1')
        self.assertIn('只能选一种', self._html(resp))

    def test_invalid_from_addr_rejected(self):
        """发件箱地址缺 @ 应被拒。"""
        resp = self._save(from_addr='not-an-email')
        self.assertIn('发件箱地址格式不正确', self._html(resp))

    def test_batch_limit_range(self):
        """单轮发送上限越界应被拒。"""
        self.assertIn('单轮发送上限', self._html(self._save(batch_limit='0')))

    def test_retry_max_range(self):
        """重试次数越界应被拒。"""
        self.assertIn('重试次数', self._html(self._save(retry_max='9')))

    def test_manual_cooldown_range(self):
        """手动冷却越界应被拒。"""
        self.assertIn('手动发送冷却', self._html(self._save(manual_cooldown='9999')))

    def test_env_locked_hint(self):
        """被 .env 锁定的配置项要在页面上标注，避免「保存了没变化」的假故障。"""
        os.environ['MAIL_ENABLED'] = '0'
        try:
            html = self._html(self.client.get('/mail'))
            self.assertIn('已被 .env / 系统环境变量锁定', html)
            self.assertIn('enabled', html)
        finally:
            del os.environ['MAIL_ENABLED']

    # --- 未填邮箱名单 ---

    def test_users_without_email_listed(self):
        """未填邮箱的用户应出现在名单里（D4-②）。"""
        html = self._html(self.client.get('/mail'))
        self.assertIn('位用户未填写邮箱', html)
        self.assertIn('张三', html)

    def test_admin_can_fill_user_email(self):
        """管理员可代填邮箱，填完即从名单移出。"""
        self.client.post('/mail/users/%d/email' % self.owner_id,
                         data={'email': 'zhangsan@example.com'},
                         follow_redirects=True)
        self.assertEqual(models.get_user(self.owner_id)['email'],
                         'zhangsan@example.com')
        left = [u['user_id'] for u in models.get_users_without_email()]
        self.assertNotIn(self.owner_id, left)

    def test_fill_user_email_validates_format(self):
        """代填邮箱需过基本格式校验。"""
        self.client.post('/mail/users/%d/email' % self.owner_id,
                         data={'email': 'bad'}, follow_redirects=True)
        self.assertIsNone(models.get_user(self.owner_id)['email'])

    # --- 界面入口与展示细节 ---

    def test_nav_mail_entry_admin_only(self):
        """导航栏「邮件」入口只对管理员渲染（P-1）。

        用 href 而不是「邮件」二字做断言：状态页里「邮件」出现十几次，
        拿文字判断等于没验证。
        """
        self.assertIn('href="/mail"', self._html(self.client.get('/tasks')))

        self.client.get('/logout')
        self._login('owner', 'owner123456')
        self.assertNotIn('href="/mail"', self._html(self.client.get('/tasks')))

    def test_configured_shows_enabled_badge(self):
        """配置齐全后横幅撤下、徽标转为「已启用」。"""
        self.assertIn('未启用', self._html(self.client.get('/mail')))
        self._save()
        html = self._html(self.client.get('/mail'))
        self.assertNotIn('邮件功能当前不可用', html)
        self.assertIn('已启用', html)

    def test_profile_page_shows_mail_section(self):
        """个人设置页要能看到邮箱、订阅等级与「我的邮件记录」入口。"""
        html = self._html(self.client.get('/settings/profile'))
        self.assertIn('邮件通知', html)
        self.assertIn('查看我的邮件记录', html)
        self.assertIn('mail_notify_level', html)
        for label in mail_constants.NOTIFY_LEVEL_LABELS.values():
            self.assertIn(label, html)

    def _log_card_count(self, html):
        """从「发送记录（共 N 条）」卡标题里取出条数。

        不能直接用「某主题是否出现」判断筛选是否生效——失败件不管怎么筛
        都会出现在上方独立的「发送失败清单」卡里，按文字断言永远为真。
        """
        m = re.search(r'发送记录（共\s*(\d+)\s*条）', html)
        return int(m.group(1)) if m else -1

    def test_logs_filter_by_status(self):
        """按成功/失败结果筛选应命中（关键字与类型之外最常用的维度）。"""
        self._log(self.admin_id, 'ok@example.com',
                  mail_constants.MAIL_TYPE_TEST, '成功的一封')
        self._log(self.admin_id, 'bad@example.com',
                  mail_constants.MAIL_TYPE_TEST, '失败的一封',
                  success=0, error='550')

        self.assertEqual(self._log_card_count(self._html(
            self.client.get('/mail'))), 2)

        ok_html = self._html(self.client.get('/mail?status=success'))
        self.assertEqual(self._log_card_count(ok_html), 1)
        self.assertIn('成功的一封', ok_html)

        bad_html = self._html(self.client.get('/mail?status=failed'))
        self.assertEqual(self._log_card_count(bad_html), 1)
        self.assertIn('失败的一封', bad_html)

    # --- 熔断 ---

    def test_circuit_open_shows_banner_and_resume(self):
        """熔断时要显示原因与恢复入口（G3-③：认证失败需人工恢复）。

        先配置好邮件再熔断：横幅是「未配置 → 熔断」的 if/elif 结构，
        未配置时顶部要先讲清「功能压根没开」，否则管理员会先去
        排查一个根本不会发送的系统为什么熔断了。
        """
        self._save()
        models.set_circuit_state('open', reason='SMTP 认证失败：535', fail_streak=3)
        html = self._html(self.client.get('/mail'))
        self.assertIn('邮件发送已熔断', html)
        self.assertIn('SMTP 认证失败：535', html)
        self.assertIn('需人工恢复', html)

        self.client.post('/mail/circuit/resume', follow_redirects=True)
        self.assertEqual(models.get_circuit_state()['state'], 'closed')

    def test_unconfigured_takes_precedence_over_circuit(self):
        """未配置时顶部横幅讲「未启用」，但熔断详情卡仍要渲染。

        管理员可能在熔断后关掉了邮件功能，再回来时如果只看得到
        「未启用」，就会以为熔断已经解除，重新启用后才发现还在暂停。
        """
        models.set_circuit_state('open', reason='SMTP 认证失败：535', fail_streak=3)
        html = self._html(self.client.get('/mail'))
        self.assertIn('邮件功能当前不可用', html)
        self.assertNotIn('邮件发送已熔断', html)
        self.assertIn('SMTP 认证失败：535', html)   # 详情卡仍在

    # --- 发送记录 ---

    def test_logs_filter_by_type_and_keyword(self):
        """按类型 + 关键字筛选应命中。"""
        self._log(self.admin_id, 'a@example.com',
                  mail_constants.MAIL_TYPE_TEST, '[督办系统] 测试邮件：配置验证')
        self._log(self.admin_id, 'b@example.com',
                  mail_constants.MAIL_TYPE_OVERDUE, '[督办] 逾期提醒：某任务')

        html = self._html(self.client.get('/mail?mail_type=test&keyword=测试'))
        self.assertIn('测试邮件：配置验证', html)
        self.assertNotIn('逾期提醒：某任务', html)

    def test_logs_pagination_out_of_range(self):
        """越界/非法页码应回退到有效页，不能 500。"""
        self.assertEqual(self.client.get('/mail?page=99').status_code, 200)
        self.assertEqual(self.client.get('/mail?page=abc').status_code, 200)

    # --- 失败清单与重发 ---

    def test_failed_list_and_requeue(self):
        """失败邮件出现在清单里，一键重发后回到队列（G7-①）。"""
        log_id = self._log(self.owner_id, 'zhangsan@example.com',
                           mail_constants.MAIL_TYPE_OVERDUE, '[督办] 逾期提醒',
                           success=0, error='550 收件人不存在')
        html = self._html(self.client.get('/mail'))
        self.assertIn('发送失败清单', html)
        self.assertIn('550 收件人不存在', html)

        resp = self.client.post('/mail/log/%d/requeue' % log_id,
                                follow_redirects=True)
        self.assertIn('已重新加入发送队列', self._html(resp))
        self.assertGreaterEqual(models.count_pending_emails(), 1)

    def test_requeue_success_rejected(self):
        """已成功的邮件不允许重发，避免重复骚扰收件人。"""
        log_id = self._log(self.admin_id, 'a@example.com',
                           mail_constants.MAIL_TYPE_TEST, '[督办系统] 测试邮件')
        resp = self.client.post('/mail/log/%d/requeue' % log_id,
                                follow_redirects=True)
        self.assertIn('已发送成功', self._html(resp))

    # --- 运维操作 ---

    def test_scan_empty_queue(self):
        """队列为空时立即扫描应给出明确提示，而不是静默无反馈。"""
        self._save()
        models.set_circuit_state('closed')
        resp = self.client.post('/mail/scan', follow_redirects=True)
        self.assertIn('队列中没有待发送的邮件', self._html(resp))

    def test_send_test_mail_without_config(self):
        """未配置时点测试邮件应给出可执行的提示，不抛异常（B5-① 静默降级）。"""
        resp = self.client.post('/mail/test', follow_redirects=True)
        self.assertIn('未启用', self._html(resp))

    # --- 普通用户视图 ---

    def test_my_mails_requires_login(self):
        """未登录访问「我的邮件记录」应跳转登录。"""
        self.client.get('/logout')
        resp = self.client.get('/mail/my')
        self.assertEqual(resp.status_code, 302)

    def test_my_mails_shows_only_mine(self):
        """H2-①：普通用户只能看到发给自己的，看不到别人的。"""
        self._log(self.admin_id, 'admin@example.com',
                  mail_constants.MAIL_TYPE_TEST, '[督办系统] 发给管理员的测试邮件')
        self._log(self.owner_id, 'zhangsan@example.com',
                  mail_constants.MAIL_TYPE_TEST, '[督办系统] 发给张三的测试邮件')

        html = self._html(self.client.get('/mail/my'))
        self.assertIn('发给管理员的测试邮件', html)
        self.assertNotIn('发给张三的测试邮件', html)

    def test_save_email_preference(self):
        """个人设置可保存邮箱与订阅等级。"""
        resp = self.client.post('/settings/mail', data={
            'email': 'admin@example.com',
            'mail_notify_level': mail_constants.LEVEL_OVERDUE_DUE,
        }, follow_redirects=True)
        self.assertIn('邮件通知设置已保存', self._html(resp))
        self.assertEqual(models.get_user(self.admin_id)['email'], 'admin@example.com')
        self.assertEqual(models.get_user(self.admin_id)['mail_notify_level'],
                         mail_constants.LEVEL_OVERDUE_DUE)

    def test_save_email_preference_validates(self):
        """个人邮箱格式非法应被拒，且不改动已有值。"""
        models.update_user_email(self.admin_id, 'admin@example.com')
        self.client.post('/settings/mail',
                         data={'email': 'bad', 'mail_notify_level': 'overdue'},
                         follow_redirects=True)
        self.assertEqual(models.get_user(self.admin_id)['email'], 'admin@example.com')

    def test_crypto_unavailable_message(self):
        """加密不可用时保存密码应给出清晰提示，而不是静默失败。"""
        real = crypto_util._derive_master_key
        try:
            crypto_util._derive_master_key = lambda: None
            resp = self._save()
            self.assertIn('加密不可用', self._html(resp))
        finally:
            crypto_util._derive_master_key = real


# ============================================================
# 21. V4 手动发送提醒邮件（任务详情页按钮）
# ============================================================

class TestManualMailSend(unittest.TestCase):
    """V4 手动发送：按钮渲染权限、冷却、AJAX 与整页两种响应形态。"""

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()
        self.client = make_client(self.app)
        self.admin_id = create_test_user('admin', '管理员',
                                         password='admin123456', role='admin')
        self.creator_id = create_test_user('creator', '布置人',
                                           password='creator123456', role='owner')
        self.owner_id = create_test_user('owner', '张三',
                                         password='owner123456', role='owner')
        # 由 creator 创建、张三负责：用于验证「创建人也能催办」
        self.task_id = create_task_direct('待催办任务', self.creator_id, self.owner_id)
        models.update_user_email(self.owner_id, 'zhangsan@example.com')

    def _login(self, username, password):
        return self.client.post(
            '/login', data={'username': username, 'password': password},
            follow_redirects=False)

    def _html(self, resp):
        return resp.data.decode('utf-8')

    def _configure_mail(self):
        """开启邮件并指向本机 1 号端口（连接会被立刻拒绝，不产生真实外发）。"""
        for key, value in MAIL_CFG_FORM.items():
            models.set_config('mail_' + key, value)
        models.set_config('mail_smtp_password',
                          crypto_util.encrypt('smtp-pass-123'))
        models.set_circuit_state('closed')

    # --- 按钮渲染 ---

    def test_admin_sees_mail_block_in_drawer(self):
        """管理员在抽屉里应看到邮件提醒区块。"""
        self._configure_mail()
        self._login('admin', 'admin123456')
        html = self._html(self.client.get('/tasks/%d/drawer' % self.task_id))
        self.assertIn('邮件提醒', html)
        self.assertIn('发送提醒邮件给负责人', html)

    def test_creator_sees_mail_block(self):
        """H1-②：任务创建人也可以催办。"""
        self._configure_mail()
        self._login('creator', 'creator123456')
        html = self._html(self.client.get('/tasks/%d/drawer' % self.task_id))
        self.assertIn('发送提醒邮件给负责人', html)

    def test_plain_owner_does_not_see_mail_block(self):
        """H1-②：既非管理员也非创建人的负责人不应看到按钮（不能自己催自己）。"""
        self._configure_mail()
        self._login('owner', 'owner123456')
        html = self._html(self.client.get('/tasks/%d/drawer' % self.task_id))
        self.assertNotIn('发送提醒邮件给负责人', html)

    def test_unconfigured_shows_hint_not_button(self):
        """未配置邮件时按钮位置显示说明文案，而不是消失（让操作人知道有这功能）。"""
        self._login('admin', 'admin123456')
        html = self._html(self.client.get('/tasks/%d/drawer' % self.task_id))
        self.assertIn('邮件提醒', html)
        self.assertNotIn('发送提醒邮件给负责人', html)
        self.assertIn('未启用', html)

    def test_detail_page_shows_mail_block(self):
        """整页详情页同样提供邮件提醒入口。"""
        self._configure_mail()
        self._login('admin', 'admin123456')
        html = self._html(self.client.get('/tasks/%d' % self.task_id))
        self.assertIn('发送提醒邮件给负责人', html)

    # --- 发送行为 ---

    def test_ajax_send_returns_json(self):
        """AJAX 提交返回 JSON，由前端弹 Toast。"""
        self._configure_mail()
        self._login('admin', 'admin123456')
        resp = self.client.post(
            '/tasks/%d/send-mail' % self.task_id, json={'note': '请尽快反馈'},
            headers={'X-Requested-With': 'XMLHttpRequest'})
        self.assertTrue(resp.is_json)
        body = resp.get_json()
        self.assertIn('success', body)
        self.assertIn('message', body)

    def test_fullpage_send_redirects_with_flash(self):
        """整页表单提交应重定向 + flash，不能返回一屏 JSON。"""
        self._configure_mail()
        self._login('admin', 'admin123456')
        resp = self.client.post('/tasks/%d/send-mail' % self.task_id,
                                data={'note': '请尽快反馈'},
                                follow_redirects=False)
        self.assertEqual(resp.status_code, 302)

    def test_send_failure_returns_202_and_keeps_in_queue(self):
        """发送失败（连不上 SMTP）时返回 202 并把记录留在队列里等后台重试。"""
        self._configure_mail()
        self._login('admin', 'admin123456')
        resp = self.client.post('/tasks/%d/send-mail' % self.task_id,
                                json={'note': 'x'},
                                headers={'X-Requested-With': 'XMLHttpRequest'})
        self.assertEqual(resp.status_code, 202)
        self.assertIn('已加入发送队列', resp.get_json()['message'])
        self.assertGreaterEqual(models.count_pending_emails(), 1)

    def test_send_forbidden_for_non_creator(self):
        """非管理员且非创建人调用接口应 403。"""
        self._configure_mail()
        self._login('owner', 'owner123456')
        resp = self.client.post('/tasks/%d/send-mail' % self.task_id,
                                json={'note': 'x'},
                                headers={'X-Requested-With': 'XMLHttpRequest'})
        self.assertEqual(resp.status_code, 403)

    def test_send_requires_login(self):
        """未登录调用发送接口应跳转登录。"""
        self._configure_mail()
        resp = self.client.post('/tasks/%d/send-mail' % self.task_id, json={})
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp.headers.get('Location', ''))

    def test_send_rejected_when_mail_unconfigured(self):
        """邮件未配置时发送接口应明确拒绝，而不是入队后永远发不出去。"""
        self._login('admin', 'admin123456')
        resp = self.client.post('/tasks/%d/send-mail' % self.task_id,
                                json={'note': 'x'},
                                headers={'X-Requested-With': 'XMLHttpRequest'})
        self.assertEqual(resp.status_code, 400)

    def test_cooldown_blocks_repeat_send(self):
        """F4-②：冷却窗口内重复发送应被拒，防止手抖刷屏。"""
        self._configure_mail()
        self._login('admin', 'admin123456')
        url = '/tasks/%d/send-mail' % self.task_id

        first = self.client.post(url, json={'note': '第一次'},
                                 headers={'X-Requested-With': 'XMLHttpRequest'})
        self.assertEqual(first.status_code, 202)   # 连不上 SMTP，入队待重试

        second = self.client.post(url, json={'note': '第二次'},
                                  headers={'X-Requested-With': 'XMLHttpRequest'})
        self.assertEqual(second.status_code, 400)
        self.assertIn('发送过于频繁', second.get_json()['message'])

    def test_cooldown_disables_button(self):
        """冷却期内抽屉里的按钮应被禁用并给出说明。"""
        self._configure_mail()
        self._login('admin', 'admin123456')
        self.client.post('/tasks/%d/send-mail' % self.task_id, json={'note': 'x'},
                         headers={'X-Requested-With': 'XMLHttpRequest'})
        html = self._html(self.client.get('/tasks/%d/drawer' % self.task_id))
        self.assertIn('冷却中', html)
        self.assertIn('disabled', html)


class TestAIDraftReminder(unittest.TestCase):
    """V5 Phase 1 ②：催办话术「页面内预填 + 可编辑 + 确认发出」。

    覆盖：入口渲染权限（管理员 + 已启用 AI）、owner 不可见、生成后页面内
    可编辑预填、确认采纳把（可编辑后的）内容作为站内信发出、生成失败回显。
    """

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()
        # 清掉 AI 队列/历史，避免跨用例残留
        conn = db.get_db()
        conn.execute('DELETE FROM ai_queue')
        conn.execute('DELETE FROM ai_log')
        conn.commit()
        self.client = make_client(self.app)
        self._orig_ai = config.AI_ENABLED
        self._orig_call = ai_service.call_model
        self.admin_id = create_test_user('admin', '管理员',
                                         password='admin123456', role='admin')
        self.creator_id = create_test_user('creator', '布置人',
                                           password='creator123456', role='owner')
        self.owner_id = create_test_user('owner', '张三',
                                         password='owner123456', role='owner')
        self.task_id = create_task_direct('待催办任务', self.creator_id, self.owner_id)

    def tearDown(self):
        # 还原 AI 开关与模型调用，避免污染其它用例
        config.AI_ENABLED = self._orig_ai
        ai_service.call_model = self._orig_call

    def _login(self, username, password):
        return self.client.post(
            '/login', data={'username': username, 'password': password},
            follow_redirects=False)

    def _html(self, resp):
        return resp.data.decode('utf-8')

    def _enable_ai(self, text='请尽快推进任务X的进展，本周内反馈'):
        """开启 AI 并把模型调用替换成可控的假响应。"""
        config.AI_ENABLED = True
        ai_service.call_model = lambda prompt: {
            'success': True, 'text': text, 'error': None}

    # --- 入口渲染权限 ---

    def test_admin_sees_ai_block_when_enabled(self):
        """AI 启用且为管理员时，详情页应出现「AI 催办话术」入口。"""
        self._enable_ai()
        self._login('admin', 'admin123456')
        html = self._html(self.client.get('/tasks/%d' % self.task_id))
        self.assertIn('AI 催办话术', html)
        self.assertIn('生成催办话术', html)

    def test_ai_block_hidden_when_disabled(self):
        """AI 未启用时，即使是管理员也不应看到入口。"""
        self._login('admin', 'admin123456')
        html = self._html(self.client.get('/tasks/%d' % self.task_id))
        self.assertNotIn('AI 催办话术', html)

    def test_owner_does_not_see_ai_block(self):
        """AI 仅管理员可用，负责人（owner）看不到入口。"""
        self._enable_ai()
        self._login('owner', 'owner123456')
        html = self._html(self.client.get('/tasks/%d' % self.task_id))
        self.assertNotIn('AI 催办话术', html)

    # --- 生成 → 预填 → 确认发出 ---

    def test_generate_prefills_editable_textarea(self):
        """点击生成后应同步生成并在详情页预填可编辑文本域。"""
        self._enable_ai(text='请于周五前反馈任务进展')
        self._login('admin', 'admin123456')
        resp = self.client.post(
            '/ai/trigger',
            data={'task_id': self.task_id, 'source': 'detail'},
            follow_redirects=True)
        html = self._html(resp)
        self.assertIn('请于周五前反馈任务进展', html)
        self.assertIn('确认发送给负责人', html)
        # 内容落在可编辑的 textarea（name=content）里
        self.assertIn('name="content"', html)

    def test_adopt_sends_edited_content_as_message(self):
        """确认采纳应把（可编辑后的）内容作为站内信发给负责人。"""
        self._enable_ai(text='原稿提醒内容')
        self._login('admin', 'admin123456')
        self.client.post(
            '/ai/trigger',
            data={'task_id': self.task_id, 'source': 'detail'},
            follow_redirects=True)
        log = models.list_ai_logs(1)[0]
        self.assertTrue(log['success'])

        edited = '人工修改后的催办话术'
        self.client.post('/ai/adopt/%d' % log['log_id'],
                         data={'content': edited, 'next': 'detail'},
                         follow_redirects=True)

        # 负责人应收到一条 ai_reminder 站内信，内容为修改后的版本
        msgs = models.get_messages(self.owner_id)
        self.assertTrue(any(m['type'] == 'ai_reminder'
                            and m['content'] == edited for m in msgs))
        self.assertTrue(models.get_ai_log(log['log_id'])['adopted'])

    def test_adopt_rejects_empty_content(self):
        """内容被清空时确认应被拒，不发出空站内信。"""
        self._enable_ai(text='原稿')
        self._login('admin', 'admin123456')
        self.client.post(
            '/ai/trigger',
            data={'task_id': self.task_id, 'source': 'detail'},
            follow_redirects=True)
        log = models.list_ai_logs(1)[0]
        self.client.post('/ai/adopt/%d' % log['log_id'],
                         data={'content': '   ', 'next': 'detail'},
                         follow_redirects=True)
        msgs = models.get_messages(self.owner_id)
        self.assertFalse(any(m['type'] == 'ai_reminder' for m in msgs))
        self.assertFalse(models.get_ai_log(log['log_id'])['adopted'])

    def test_generate_failure_shows_error(self):
        """模型调用失败应在详情页回显错误，并可重新生成。"""
        config.AI_ENABLED = True
        ai_service.call_model = lambda prompt: {
            'success': False, 'text': None, 'error': '连接本地模型超时'}
        self._login('admin', 'admin123456')
        resp = self.client.post(
            '/ai/trigger',
            data={'task_id': self.task_id, 'source': 'detail'},
            follow_redirects=True)
        html = self._html(resp)
        self.assertIn('本次生成失败', html)
        self.assertIn('连接本地模型超时', html)
        self.assertIn('重新生成', html)


class TestStaticAssetVersion(unittest.TestCase):
    """静态资源缓存串（cache-busting）：防「改了样式忘了升版本号」回归。

    背景（2026-09-03 真实事故）：模板里 url_for('static', ..., v='YYYYMMDDx')
    的版本号原本写死在每个模板中。某次改 CSS 只 bump 了 base.html，而
    login.html / setup.html 是独立页面（不继承 base.html），版本号没跟着变，
    结果登录页与初始化向导页仍向浏览器请求旧版本 → 命中缓存 → 拿到旧样式，
    打包成 exe 后尤其难查（只能靠实跑冒烟发现）。

    现在版本号集中在 config.STATIC_VERSION，由 inject_globals() 注入。
    下面两个用例分别守住「结构」（不许再硬编码）与「行为」
    （每个页面真的拿到同一个版本号）。
    """

    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.app.config['TESTING'] = True

    def setUp(self):
        reset_database()
        self.client = make_client(self.app)

    def _asset_versions(self, html):
        """取出页面里所有静态资源的缓存串（去重排序）。"""
        return sorted(set(re.findall(r'/static/[^\s"\']+\?v=([^"\'\s>]+)', html)))

    def test_no_hardcoded_version_in_templates(self):
        """模板里不许出现 v='YYYYMMDD' 字面量，一律写 v=static_version。

        谁把版本号写回某个模板，这个用例就失败——那意味着下次改样式时
        又会漏掉这个页面（login.html / setup.html 不继承 base.html，最易漏）。
        """
        offenders = []
        pattern = re.compile(r"url_for\(\s*'static'[^)]*v\s*=\s*'[^']*'")
        for dirpath, _dirnames, filenames in os.walk(config.TEMPLATE_DIR):
            for name in filenames:
                if not name.endswith('.html'):
                    continue
                path = os.path.join(dirpath, name)
                with open(path, encoding='utf-8') as f:
                    for lineno, line in enumerate(f, 1):
                        if pattern.search(line):
                            rel = os.path.relpath(path, config.TEMPLATE_DIR)
                            offenders.append('%s:%d' % (rel, lineno))
        self.assertEqual([], offenders,
                         '静态资源版本号必须写成 v=static_version（由 config.STATIC_VERSION '
                         '统一提供），以下位置仍在硬编码：%s' % '、'.join(offenders))

    def test_every_page_uses_the_same_version(self):
        """登录页 / 初始化向导页 / 已登录页面必须拿到同一个版本号。

        不打真实页面就测不出来——模板源码看着对，渲染时上下文没注入变量
        同样会渲染成空串。
        """
        expected = config.STATIC_VERSION
        self.assertTrue(expected, 'config.STATIC_VERSION 不能为空')

        # 初始化向导页：必须在建号之前取——库里一旦有管理员，
        # /setup 就会跳回登录页，同样 /login 也会跳去 /setup，
        # 拿到的都是 302 跳转存根而不是真正的页面。
        setup_html = self.client.get('/setup').data.decode('utf-8')
        self.assertEqual([expected], self._asset_versions(setup_html),
                         '初始化向导页未使用统一版本号')

        create_test_user('admin', '管理员', password='admin123456', role='admin')

        # 登录页
        login_html = self.client.get('/login').data.decode('utf-8')
        self.assertEqual([expected], self._asset_versions(login_html),
                         '登录页未使用统一版本号——改了 CSS 这里会拿到旧样式')

        # 已登录页面（走 base.html）
        self.client.post('/login', data={'username': 'admin', 'password': 'admin123456'})
        dash_html = self.client.get('/dashboard').data.decode('utf-8')
        self.assertEqual([expected], self._asset_versions(dash_html),
                         '已登录页面未使用统一版本号')


# ============================================================
# 主程序入口
# ============================================================

if __name__ == '__main__':
    print('=' * 60)
    print('督办系统综合测试套件')
    print('=' * 60)
    print()

    # 创建测试套件
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()

    suite.addTests(loader.loadTestsFromTestCase(TestDatabaseInit))
    suite.addTests(loader.loadTestsFromTestCase(TestStateMachine))
    suite.addTests(loader.loadTestsFromTestCase(TestAppStartup))
    suite.addTests(loader.loadTestsFromTestCase(TestEndToEnd))
    suite.addTests(loader.loadTestsFromTestCase(TestPermissions))
    suite.addTests(loader.loadTestsFromTestCase(TestPagination))
    suite.addTests(loader.loadTestsFromTestCase(TestWarningEngine))
    suite.addTests(loader.loadTestsFromTestCase(TestDashboard))
    suite.addTests(loader.loadTestsFromTestCase(TestBatchAndExport))
    # V2 新增用例
    suite.addTests(loader.loadTestsFromTestCase(TestDashboardV2))
    suite.addTests(loader.loadTestsFromTestCase(TestInlineFieldEdit))
    suite.addTests(loader.loadTestsFromTestCase(TestEvidenceBlockers))
    suite.addTests(loader.loadTestsFromTestCase(TestTimelineLogging))
    suite.addTests(loader.loadTestsFromTestCase(TestDrawerRoutes))
    suite.addTests(loader.loadTestsFromTestCase(TestRemind))
    suite.addTests(loader.loadTestsFromTestCase(TestLoginV2))
    suite.addTests(loader.loadTestsFromTestCase(TestTaskFormExtraFields))
    # DEF-005：登录限流与弱口令校验
    suite.addTests(loader.loadTestsFromTestCase(TestLoginSecurity))
    suite.addTests(loader.loadTestsFromTestCase(TestBehindProxy))
    # DEF-002：CSRF 防护
    suite.addTests(loader.loadTestsFromTestCase(TestCSRF))
    # V5 AI 辅助生成（Phase 1 ②）
    suite.addTests(loader.loadTestsFromTestCase(TestAIDraftReminder))

    # 运行测试
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)

    print()
    print('=' * 60)
    print('测试结果汇总')
    print('=' * 60)
    print('总测试数: %d' % result.testsRun)
    print('通过: %d' % (result.testsRun - len(result.failures) - len(result.errors)))
    print('失败: %d' % len(result.failures))
    print('错误: %d' % len(result.errors))
    print()

    if result.failures:
        print('--- 失败用例 ---')
        for test, traceback in result.failures:
            print('  FAIL: %s' % test)
            print('  %s' % traceback[:500])
        print()

    if result.errors:
        print('--- 错误用例 ---')
        for test, traceback in result.errors:
            print('  ERROR: %s' % test)
            print('  %s' % traceback[:500])
        print()

    sys.exit(0 if result.wasSuccessful() else 1)
